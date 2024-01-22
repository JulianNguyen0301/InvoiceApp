import pytesseract 
from pytesseract import Output
import numpy as np
import cv2
import os
import xlwings as xw
from pdf2image import convert_from_path
import re
from datetime import datetime
from selenium import webdriver
from time import sleep,time
from selenium.webdriver.common.by import By
import img2pdf
import glob
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Font,numbers,NamedStyle,Color
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl    
from pathlib import Path
import tensorflow as tf
from tensorflow import keras
from random import *
import pandas as pd

def pdf_to_png_PO():
    poppler_path = r'H:/OCR/Popler/poppler-23.07.0/Library/bin'
    pdf_folder = r"H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR/File_PDF_PO"
    saving_folder = r"H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Compare_PO"
    os.makedirs(saving_folder, exist_ok=True)
    for pdf_filename in os.listdir(pdf_folder):
        if pdf_filename.lower().endswith(".pdf"):
            pdf_path = os.path.join(pdf_folder, pdf_filename)
            pages = convert_from_path(pdf_path=pdf_path, poppler_path=poppler_path)
            for c, page in enumerate(pages, start=1):
                img_name = f"{os.path.splitext(pdf_filename)[0]}.png"
                img_path = os.path.join(saving_folder, img_name)
                page.save(img_path, "png")
        os.remove(pdf_path)

def Excel_Jusified_PO():
    def clear_cells(sheet, start_col, end_col, row_number):
        for col in range(start_col, end_col + 1):
            sheet.cell(row=row_number, column=col).value = ''

    def format_number_with_commas(number_str):
            if number_str is None:
                return ""
            
            if isinstance(number_str, int):
                number_str = str(number_str)
            
            number_str = re.sub(r'[,]', '', number_str)  # Loại bỏ dấu phẩy
            try:
                number = int(number_str)
                formatted_number = f"{number:,}"
                return formatted_number
            except ValueError:
                return number_str

    def convert_string_to_number(s):
            try:
                s = s.replace(',', '')  # Loại bỏ dấu phẩy (,) trong chuỗi
                return float(s)  # Chuyển đổi thành số thực
            except ValueError:
                return None  # Không thể chuyển đổi thành số
    location = 'H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_PO\\*.xlsx'
    excel_files = glob.glob(location)
    output_path = 'H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_PO\\Total_PO.xlsx'

    workbook = Workbook()
    for excel_file in excel_files:
        sheet_name = os.path.basename(excel_file)[:31]
        df1 = pd.read_excel(excel_file, engine='openpyxl')
        df1.fillna(value='', inplace=True)
        sheet = workbook.create_sheet(sheet_name)
        for row in dataframe_to_rows(df1, index=False, header=True):
            sheet.append(row)
    # Xóa trang mặc định
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)
    # Lưu Workbook vào tệp mới
    workbook.save(output_path)
    output_excel_file = 'H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_PO\\Total_PO.xlsx'
    excel = openpyxl.load_workbook(output_excel_file)
    sheet_names = excel.sheetnames
    font = Font(name='Times New Roman', size=12)
    sheet_counter = 0 
    for sheet_name in sheet_names:
        sheet = excel[sheet_name]
        max_widths = []
        clear_cells(sheet, 3, 4, 1) 
        for row in sheet.iter_rows():
            for i, cell in enumerate(row):
                cell.font = font
                cell_value = str(cell.value)
                if i >= len(max_widths):
                    max_widths.append(len(cell_value))
                else:
                    max_widths[i] = max(max_widths[i], len(cell_value))
        for i, column_width in enumerate(max_widths, start=1):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = column_width + 2
        for row in sheet.iter_rows(min_row=1, min_col=1, max_col=len(max_widths)):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
        for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
        for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='right', vertical='center')
        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='right', vertical='center')
        for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')    
        red_font = Font(color=Color(rgb="FF0000"), bold=True,name='Times New Roman', size=12)
        sheet['F3'].font = red_font
        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)
                
        for row in sheet.iter_rows(min_row=8, max_row=8, min_col=1, max_col=7):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)

        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=5, max_col=5):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)
        for row in sheet.iter_rows(min_row=4, max_row=6, min_col=6, max_col=6):
            for cell in row:
                cell_value = format_number_with_commas(cell.value)
                cell.value = cell_value
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, min_col=5, max_col=6):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell_value = format_number_with_commas(str(cell.value))
                    cell.value = cell_value
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sum = 0
        for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, min_col=6, max_col=6):
            for cell in row:
                cell_value = str(cell.value)
                number_value = convert_string_to_number(cell_value)
                
                if number_value is not None:
                    sum += number_value
        cell_F4 = sheet['F4']
        value_F4 = convert_string_to_number(str(cell_F4.value))
        if sum == value_F4:
            None
        else:
            print(f"Sheet {sheet_counter + 1} có tổng thành tiền khác với cộng tiền hàng chưa thuế.")
        sheet_counter += 1
    # Lưu lại tệp Excel đã cập nhật
    output_filename_with_time = f'Total_PO_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    output_excel_file_with_time = os.path.join("H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_PO", output_filename_with_time)
    excel.save(output_excel_file_with_time)
    total_PO_path = os.path.join("H:/APP UNIVERSITY/CODE PYTHON/OpenCV-Master-Computer-Vision-in-Python/SourcecodeOCR/Data_PO", "Total_PO.xlsx")
    os.remove(total_PO_path)

    excel = openpyxl.load_workbook(output_excel_file_with_time)
    new_sheet_name = "Data_Total"
    new_sheet = excel.create_sheet(new_sheet_name)
    new_sheet['A12'] = "STT"
    new_sheet['B12'] = "Ký hiệu PO"
    new_sheet['C12'] = "Ngày giao hàng"
    new_sheet['D12'] = "Tên người bán"
    new_sheet['E12'] = "Mặt hàng"
    new_sheet['F12'] = "Doanh số mua chưa có thuế"
    new_sheet['G12'] = "Thuế GTGT đủ điều kiện khấu trừ thuế"
    new_sheet['H12'] = "Ghi chú"
    for column in new_sheet.iter_cols(min_col=1, max_col=8, min_row=12, max_row=12):
        for cell in column:
            cell.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')
            cell.font = Font(name='Times New Roman', size=12,bold=True)
    column_widths = {'A': 5,'B': 10,'C': 10,'D': 75,'E': 40,'F': 15,'G': 15,'H': 50}
    for column, width in column_widths.items():
        new_sheet.column_dimensions[column].width = width

    excel.save(output_excel_file_with_time)
    count_sheet = 0
    sheets_with_max_row_9 = []
    for sheet_name in excel.sheetnames[:-1]:  
        sheet = excel[sheet_name]
        new_row = len(new_sheet['A'])+1  # Bắt đầu từ hàng 13
        data_to_copy = sheet['F1'].value
        new_sheet.cell(row=new_row, column=2, value=data_to_copy)
        data_to_copy = sheet['B6'].value
        new_sheet.cell(row=new_row, column=3, value=data_to_copy)
        data_to_copy = sheet['B3'].value
        new_sheet.cell(row=new_row, column=4, value=data_to_copy)
        data_to_copy = sheet['F4'].value
        new_sheet.cell(row=new_row, column=6, value=data_to_copy)
        data_to_copy = sheet['F5'].value
        new_sheet.cell(row=new_row, column=7, value=data_to_copy)
        count_sheet +=1
        if sheet['F3'].value == "Valid":
            data_to_copy = sheet['F3'].value
            new_sheet.cell(row=new_row, column=8, value=data_to_copy)
        elif sheet['F3'].value == "Invalid":
            data_to_copy = sheet['G1'].value
            new_sheet.cell(row=new_row, column=8, value=data_to_copy)
            sheet['G1'].value = ''
        elif sheet['F3'].value == "":
            shd = sheet['F1'].value
            data_to_copy = f"Không có PO để so sánh với Invoice có số hóa đơn là {shd}"
            new_sheet.cell(row=new_row, column=8, value=data_to_copy)
        if sheet.max_row == 9:
            sheets_with_max_row_9.append(sheet_name)
            count_note = count_sheet
            if sheets_with_max_row_9:
                target_sheet_name = sheets_with_max_row_9[0] 
                target_worksheet = excel[target_sheet_name]
                value_at_B9 = target_worksheet['B19'].value
                last_worksheet_name = excel.sheetnames[-1]
                last_worksheet = excel[last_worksheet_name]
                last_worksheet[f'G{12 + count_note}'] = value_at_B9
                excel.save(output_excel_file_with_time)
                sheets_with_max_row_9.pop(0)

    for row in range(13, new_sheet.max_row + 1):
        cell_value = new_sheet.cell(row=row, column=5).value
        if ((new_sheet.cell(row=row, column=4).value == "CÔNG TY TRÁCH NHIỆM HỮU HẠN NƯỚC GIẢI KHÁT COCA-COLA VIỆT NAM") and (cell_value == None)) or ((new_sheet.cell(row=row, column=4).value == "CHI NHÁNH CÔNG TY TNHH NƯỚC GIẢI KHÁT COCA-COLA VIỆT NAM TẠI THÀNH PHỐ ĐÀ NẴNG") and (cell_value == None)):
            new_sheet.cell(row=row, column=5, value="Chi phí mua đồ uống giải khát")
        # elif new_sheet.cell(row=row, column=4).value == "CÔNG TY TNHH THỰC PHẨM LINH KHOA" and cell_value == None:
        #     new_sheet.cell(row=row, column=5, value="Chi phí mua thực phẩm")

    for column in new_sheet.iter_cols(min_col=1, max_col=8, min_row=13, max_row=new_sheet.max_row):
        for cell in column:
            cell.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')
    for index, number in enumerate(range(1, count_sheet + 1)):
        destination_row = 13 + index
        new_sheet.cell(row=destination_row, column=1, value=number)
   
    number_style = NamedStyle(name='number_style', number_format='#,##0')
    for row in new_sheet.iter_rows(min_row=13, max_row=new_sheet.max_row, min_col=6, max_col=7):
        for cell in row:
            cell.style = number_style
    for column in new_sheet.iter_cols(min_col=6, max_col=7, min_row=13, max_row=new_sheet.max_row):
        for cell in column:
            cell.alignment = Alignment(wrapText=True,horizontal='right', vertical='center')
    for column in new_sheet.iter_cols(min_col=1, max_col=8, min_row=13, max_row=new_sheet.max_row):
        for cell in column:
            cell.font = Font(name='Times New Roman', size=11)
    output_folder = "H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Data_PO_Final"
    output_excel_file_with_time1 = os.path.join(output_folder, output_filename_with_time)
    excel.save(output_excel_file_with_time1)
    workbook = load_workbook(output_excel_file_with_time1)
    worksheet_names = workbook.sheetnames
    new_names = [f"PO{i+1}" for i in range(len(worksheet_names)-1)]
    for i in range(len(worksheet_names)-1):
        worksheet = workbook[worksheet_names[i]]
        worksheet.title = new_names[i]
    workbook.save(output_excel_file_with_time1)
    os.remove(output_excel_file_with_time)
    path_po = "H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Data_PO"
    for filename in os.listdir(path_po):
        file_path = os.path.join(path_po, filename)
        os.remove(file_path)
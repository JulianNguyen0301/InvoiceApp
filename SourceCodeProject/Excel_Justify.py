import pandas as pd
import glob
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Font,numbers,NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl    
import re
import datetime
def Excel_Jusified():
    def clear_cells(sheet, start_col, end_col, row_number):
        for col in range(start_col, end_col + 1):
            sheet.cell(row=row_number, column=col).value = ''

    def format_number_with_commas(number_str):
        if number_str is None:
            return ""
        
        if isinstance(number_str, int):
            number_str = str(number_str)
        
        number_str = re.sub(r'[,]', '', number_str)  # Loại bỏ dấu phẩy
        try:
            number = int(number_str)
            formatted_number = f"{number:,}"
            return formatted_number
        except ValueError:
            return number_str
    def convert_string_to_number(s):
        try:
            s = s.replace(',', '')  # Loại bỏ dấu phẩy (,) trong chuỗi
            return float(s)  # Chuyển đổi thành số thực
        except ValueError:
            return None  # Không thể chuyển đổi thành số
    location = 'H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_Excel\\*.xlsx'
    excel_files = glob.glob(location)
    output_path = 'H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_Excel\\Total_Invoice.xlsx'

    workbook = Workbook()
    # Duyệt qua các tệp Excel để đọc dữ liệu và ghi vào tệp mới
    for excel_file in excel_files:
        sheet_name = os.path.basename(excel_file)[:31]
        df1 = pd.read_excel(excel_file, engine='openpyxl')
        df1.fillna(value='', inplace=True)
        
        # Tạo một WorkSheet mới trong Workbook
        sheet = workbook.create_sheet(sheet_name)
        
        # Ghi dữ liệu từ DataFrame vào WorkSheet
        for row in dataframe_to_rows(df1, index=False, header=True):
            sheet.append(row)

    # Xóa trang mặc định
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)

    # Lưu Workbook vào tệp mới
    workbook.save(output_path)

    output_excel_file = 'H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_Excel\\Total_Invoice.xlsx'

    excel = openpyxl.load_workbook(output_excel_file)
    sheet_names = excel.sheetnames
    font = Font(name='Times New Roman', size=12)
    sheet_counter = 0 
    for sheet_name in sheet_names:
        sheet = excel[sheet_name]
        max_widths = []
        clear_cells(sheet, 3, 7, 1) 
        for row in sheet.iter_rows():
            
            for i, cell in enumerate(row):
                cell.font = font
                cell_value = str(cell.value)
                if i >= len(max_widths):
                    max_widths.append(len(cell_value))
                else:
                    max_widths[i] = max(max_widths[i], len(cell_value))
        for i, column_width in enumerate(max_widths, start=1):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = column_width + 2
        for row in sheet.iter_rows(min_row=1, min_col=1, max_col=len(max_widths)):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        for row in sheet.iter_rows(min_row=1, max_row=8, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
        for row in sheet.iter_rows(min_row=10, max_row=25, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
        for row in sheet.iter_rows(min_row=10, max_row=25, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='right', vertical='center')
        for row in sheet.iter_rows(min_row=4, max_row=7, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='right', vertical='center')
        for row in sheet.iter_rows(min_row=10, max_row=25, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')        
        for row in sheet.iter_rows(min_row=1, max_row=8, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)
                
        for row in sheet.iter_rows(min_row=9, max_row=9, min_col=1, max_col=7):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)

        for row in sheet.iter_rows(min_row=4, max_row=7, min_col=5, max_col=5):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)
        for row in sheet.iter_rows(min_row=4, max_row=7, min_col=6, max_col=6):
            for cell in row:
                cell_value = format_number_with_commas(cell.value)
                cell.value = cell_value
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        for row in sheet.iter_rows(min_row=10   , max_row=sheet.max_row, min_col=5, max_col=6):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell_value = format_number_with_commas(str(cell.value))
                    cell.value = cell_value
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sum = 0
        for row in sheet.iter_rows(min_row=10, max_row=sheet.max_row, min_col=6, max_col=6):
            for cell in row:
                cell_value = str(cell.value)
                number_value = convert_string_to_number(cell_value)
                
                if number_value is not None:
                    sum += number_value
        cell_F4 = sheet['F4']
        value_F4 = convert_string_to_number(str(cell_F4.value))
        if sum == value_F4:
            None
        else:
            print(f"Sheet {sheet_counter + 1} có tổng thành tiền khác với cộng tiền hàng chưa thuế.")
        sheet_counter += 1
    # Lưu lại tệp Excel đã cập nhật
    output_filename_with_time = f'Total_Invoice_{datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    output_excel_file_with_time = os.path.join("H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_Excel", output_filename_with_time)
    excel.save(output_excel_file_with_time)
    total_invoice_path = os.path.join("H:/APP UNIVERSITY/CODE PYTHON/OpenCV-Master-Computer-Vision-in-Python/SourcecodeOCR/Data_Excel", "Total_Invoice.xlsx")
    os.remove(total_invoice_path)

    excel = openpyxl.load_workbook(output_excel_file_with_time)
    new_sheet_name = "Dữ_Liệu_Tổng_Hợp"
    new_sheet = excel.create_sheet(new_sheet_name)
    new_sheet['A9'] = "Người nộp thuế: CÔNG TY TNHH EAST - WEST RESTAURANT CONCEPTS"
    new_sheet['A10'] = "Mã số thuế: 0071052970"
    new_sheet['A12'] = "STT"
    new_sheet['B12'] = "Ký hiệu hóa đơn"
    new_sheet['C12'] = "Số hóa đơn"
    new_sheet['D12'] = "Ngày, tháng, năm lập hóa đơn"
    new_sheet['E12'] = "Tên người bán"
    new_sheet['F12'] = "Mã số thuế người bán"
    new_sheet['G12'] = "Mặt hàng"
    new_sheet['H12'] = "Doanh số mua chưa có thuế"
    new_sheet['I12'] = "Thuế GTGT đủ điều kiện khấu trừ thuế"
    new_sheet['J12'] = "Ghi chú"
    for column in new_sheet.iter_cols(min_col=1, max_col=10, min_row=12, max_row=12):
        for cell in column:
            cell.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')
            cell.font = Font(name='Times New Roman', size=12,bold=True)
    for column in new_sheet.iter_cols(min_col=1, max_col=1, min_row=9, max_row=10):
        for cell in column:
            cell.font = Font(name='Times New Roman', size=12, bold=True)
    column_widths = {'A': 5,'B': 10,'C': 10,'D': 12,'E': 67,'F': 25,'G': 25,'H': 15,'I': 15, 'J': 15}
    for column, width in column_widths.items():
        new_sheet.column_dimensions[column].width = width

    excel.save(output_excel_file_with_time)
    count_sheet = 0
    for sheet_name in excel.sheetnames[:-1]:  
        sheet = excel[sheet_name]
        new_row = len(new_sheet['A'])+1  # Bắt đầu từ hàng 13
        data_to_copy = sheet['B5'].value
        new_sheet.cell(row=new_row, column=2, value=data_to_copy)
        data_to_copy = sheet['B6'].value
        new_sheet.cell(row=new_row, column=3, value=data_to_copy)
        data_to_copy = sheet['B7'].value
        new_sheet.cell(row=new_row, column=4, value=data_to_copy)
        data_to_copy = sheet['B1'].value
        new_sheet.cell(row=new_row, column=5, value=data_to_copy)
        data_to_copy = sheet['B3'].value
        new_sheet.cell(row=new_row, column=6, value=data_to_copy)
        data_to_copy = sheet['F4'].value
        new_sheet.cell(row=new_row, column=8, value=data_to_copy)
        data_to_copy = sheet['F5'].value
        new_sheet.cell(row=new_row, column=9, value=data_to_copy)
        count_sheet +=1
    for column in new_sheet.iter_cols(min_col=4, max_col=4, min_row=13, max_row=sheet.max_row):
        for cell in column:
            cell.alignment = Alignment(wrapText=True,horizontal='right', vertical='center')
    for column in new_sheet.iter_cols(min_col=1, max_col=1, min_row=13, max_row=sheet.max_row):
        for cell in column:
            cell.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')
    for index, number in enumerate(range(1, count_sheet + 1)):
        destination_row = 13 + index
        new_sheet.cell(row=destination_row, column=1, value=number)
    for row in range(13, new_sheet.max_row + 1):
        if new_sheet.cell(row=row, column=5).value == "CÔNG TY TRÁCH NHIỆM HỮU HẠN NƯỚC GIẢI KHÁT COCA-COLA VIỆT NAM":
            new_sheet.cell(row=row, column=7, value="Chi phí mua đồ uống giải khát")
        elif new_sheet.cell(row=row, column=5).value == "CÔNG TY TNHH THỰC PHẨM LINH KHOA":
            new_sheet.cell(row=row, column=7, value="Chi phí mua thực phẩm")
    number_style = NamedStyle(name='number_style', number_format='#,##0')
    for row in new_sheet.iter_rows(min_row=13, max_row=new_sheet.max_row, min_col=8, max_col=9):
        for cell in row:
            cell.style = number_style
    for column in new_sheet.iter_cols(min_col=8, max_col=9, min_row=13, max_row=sheet.max_row):
        for cell in column:
            cell.alignment = Alignment(wrapText=True,horizontal='right', vertical='center')
    for column in new_sheet.iter_cols(min_col=1, max_col=10, min_row=13, max_row=sheet.max_row):
        for cell in column:
            cell.font = Font(name='Times New Roman', size=10)

    output_folder = 'SourcecodeOCR/Data_Final_Excel'
    output_excel_file_with_time = os.path.join(output_folder, output_filename_with_time)
    excel.save(output_excel_file_with_time)
    folder_to_clear = 'SourcecodeOCR/Data_Excel'
    for filename in os.listdir(folder_to_clear):
        file_path = os.path.join(folder_to_clear, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)
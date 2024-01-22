import pytesseract 
from pytesseract import Output
import numpy as np
import cv2
import os
import xlwings as xw
from pdf2image import convert_from_path
import re
from datetime import datetime
from selenium import webdriver
from time import sleep,time
from selenium.webdriver.common.by import By
import img2pdf
import glob
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Font,numbers,NamedStyle,Color
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl    
from pathlib import Path
import tensorflow as tf
from tensorflow import keras
from random import *
import pandas as pd

pytesseract.pytesseract.tesseract_cmd =r'H:\\APP UNIVERSITY\\CODE PYTHON\\Tesseract-ocr\\tesseract.exe'

def predict_captcha(image_path1):
    data_dir = Path("H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\captcha_ocr_library\Capcha_img")
    images = sorted(list(map(str, list(data_dir.glob("*.png")))))          
    labels = [img.split(os.path.sep)[-1].split(".png")[0] for img in images]
    characters = set(char for label in labels for char in label)
    characters = sorted(list(characters))
    img_width = 122
    img_height = 27
    max_length = max([len(label) for label in labels])
    char_to_num = tf.keras.layers.experimental.preprocessing.StringLookup(vocabulary=list(characters), num_oov_indices=1, mask_token=None)
    num_to_char = tf.keras.layers.experimental.preprocessing.StringLookup(vocabulary=char_to_num.get_vocabulary(), mask_token=None, invert=True)

    class CTCLayer(tf.keras.layers.Layer):
        def __init__(self, name=None):
            super().__init__(name=name)
            self.loss_fn = keras.backend.ctc_batch_cost

        def call(self, y_true, y_pred):
            batch_len = tf.cast(tf.shape(y_true)[0], dtype="int64")
            input_length = tf.cast(tf.shape(y_pred)[1], dtype="int64")
            label_length = tf.cast(tf.shape(y_true)[1], dtype="int64")

            input_length = input_length * tf.ones(shape=(batch_len, 1), dtype="int64")
            label_length = label_length * tf.ones(shape=(batch_len, 1), dtype="int64")

            loss = self.loss_fn(y_true, y_pred, input_length, label_length)
            self.add_loss(loss)

            return y_pred

    my_h5_saved_model = tf.keras.models.load_model('H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\Savemodel\my_h5_saved_model.h5', custom_objects={'CTCLayer': CTCLayer})

    prediction_model = keras.models.Model(my_h5_saved_model.get_layer(name="image").input, my_h5_saved_model.get_layer(name="dense2").output)

    def decode_batch_predictions(pred):
        input_len = np.ones(pred.shape[0]) * pred.shape[1]
        results = keras.backend.ctc_decode(pred, input_length=input_len, greedy=True)[0][0][:, :max_length]
        output_text = []
        for res in results:
            res = tf.strings.reduce_join(num_to_char(res)).numpy().decode("utf-8")
            output_text.append(res)
        return output_text
    def preprocess_image(image_path):
        img = tf.io.read_file(image_path)
        img = tf.io.decode_png(img, channels=1)  
        img = tf.image.convert_image_dtype(img, tf.float32) 
        img = tf.image.resize(img, [img_height, img_width]) 
        img = tf.transpose(img, perm=[1, 0, 2]) 
        return img
    new_img = preprocess_image(image_path1)
    new_img = tf.expand_dims(new_img, axis=0)
    preds = prediction_model.predict(new_img)
    pred_texts = decode_batch_predictions(preds)
    return pred_texts[0]

def rows_bold(excel_file_path, start_column, end_column, row):
    wb = xw.Book(excel_file_path)
    sheet = wb.sheets.active
    for col in range(start_column, end_column + 1):
        sheet.range(row, col).api.Font.Bold = True
    wb.save()

def columns_bold(excel_file_path, start_row,end_row,col):
    wb = xw.Book(excel_file_path)
    sheet = wb.sheets.active
    for row in range(start_row, end_row + 1):
        sheet.range(row, col).api.Font.Bold = True
    wb.save()

def process_all_worksheets(excel_file_path):
    def fit_column_widths_for_one_sheet(sheet):
        cell_widths = {}
        used_range = sheet.used_range
        for col in range(1, used_range.columns.count + 1):
            for row in range(1, used_range.rows.count + 1):
                cell_value = used_range(row, col).value
                if cell_value:
                    if str(cell_value)[0] == "=":
                        continue
                    cell_widths[col] = max(
                        (cell_widths.get(col, 0), len(str(cell_value))+10)
                    )
        for col, column_width in cell_widths.items():
            column_width = str(column_width)
            sheet.range((1, col), (used_range.rows.count, col)).column_width = column_width

    wb = xw.Book(excel_file_path)
    for sheet in wb.sheets:
        fit_column_widths_for_one_sheet(sheet)
        sheet.range("A9:F9").api.HorizontalAlignment = -4108  # -4108 tương ứng với giá trị xlCenter trong Excel
        sheet.range("A10:A15").api.HorizontalAlignment = -4108
        sheet.range("A1:B7").api.HorizontalAlignment = -4131
        sheet.range("C10:E15").api.HorizontalAlignment = -4108
        sheet.range("F1:F4").api.HorizontalAlignment = -4152
        sheet.range("F10:F15").api.HorizontalAlignment = -4152
    wb.save()
    wb.close()

def change_font_size(wb, font_size):
    for sheet in wb.sheets:
        for row in sheet.used_range.rows:
            for cell in row:
                cell.api.Font.Size = font_size
    wb.save()

def Excel_Jusified_INV():
    def format_number_with_commas(number_str):
            if number_str is None:
                return ""
            
            if isinstance(number_str, int):
                number_str = str(number_str)
            
            number_str = re.sub(r'[,]', '', number_str)  # Loại bỏ dấu phẩy
            try:
                number = int(number_str)
                formatted_number = f"{number:,}"
                return formatted_number
            except ValueError:
                return number_str

    def convert_string_to_number(s):
            try:
                s = s.replace(',', '')  # Loại bỏ dấu phẩy (,) trong chuỗi
                return float(s)  # Chuyển đổi thành số thực
            except ValueError:
                return None  # Không thể chuyển đổi thành số
    location = 'H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_Excel\\*.xlsx'
    excel_files = glob.glob(location)
    output_path = 'H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_Excel\\Total_Invoice.xlsx'
    print()
    workbook = Workbook()
    for excel_file in excel_files:
        sheet_name = os.path.basename(excel_file)[:31]
        df1 = pd.read_excel(excel_file, engine='openpyxl')
        df1.fillna(value='', inplace=True)
        sheet = workbook.create_sheet(sheet_name)
        for row in dataframe_to_rows(df1, index=False, header=True):
            sheet.append(row)
    # Xóa trang mặc định
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)
    # Lưu Workbook vào tệp mới
    workbook.save(output_path)
    output_excel_file = 'H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_Excel\\Total_Invoice.xlsx'
    excel = openpyxl.load_workbook(output_excel_file)
    sheet_names = excel.sheetnames
    font = Font(name='Times New Roman', size=12)
    sheet_counter = 0 
    for sheet_name in sheet_names:
        sheet = excel[sheet_name]
        max_widths = []
        for row in sheet.iter_rows():
            for i, cell in enumerate(row):
                cell.font = font
                cell_value = str(cell.value)
                if i >= len(max_widths):
                    max_widths.append(len(cell_value))
                else:
                    max_widths[i] = max(max_widths[i], len(cell_value))
        for i, column_width in enumerate(max_widths, start=1):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = column_width + 2
        for row in sheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        for row in sheet.iter_rows(min_row=1, max_row=5, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in sheet.iter_rows(min_row=1, max_row=5, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')

        for row in sheet.iter_rows(min_row=1, max_row=4, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in sheet.iter_rows(min_row=1, max_row=4, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')

        for row in sheet.iter_rows(min_row=9, max_row=9, min_col=1, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')            

        for row in sheet.iter_rows(min_row=10, max_row= sheet.max_row, min_col=1, max_col=1): #sheet.max_row
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')        
        for row in sheet.iter_rows(min_row=10, max_row=sheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in sheet.iter_rows(min_row=10, max_row=sheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in sheet.iter_rows(min_row=10, max_row=sheet.max_row, min_col=3, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in sheet.iter_rows(min_row=10, max_row=sheet.max_row, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center') 

        for row in sheet.iter_rows(min_row=1, max_row=5, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')
             
        for row in sheet.iter_rows(min_row=1, max_row=4, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in sheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)
                
        for row in sheet.iter_rows(min_row=9, max_row=9, min_col=1, max_col=6):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)
        for row in sheet.iter_rows(min_row=1, max_row=5, min_col=3, max_col=3):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)
        for row in sheet.iter_rows(min_row=1, max_row=4, min_col=5, max_col=5):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)
        for row in sheet.iter_rows(min_row=1, max_row=4, min_col=6, max_col=6):
            for cell in row:
                cell_value = format_number_with_commas(cell.value)
                cell.value = cell_value
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        for row in sheet.iter_rows(min_row=10   , max_row=sheet.max_row, min_col=5, max_col=6):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell_value = format_number_with_commas(str(cell.value))
                    cell.value = cell_value
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sum = 0
        for row in sheet.iter_rows(min_row=10, max_row=sheet.max_row, min_col=6, max_col=6):
            for cell in row:
                cell_value = str(cell.value)
                number_value = convert_string_to_number(cell_value)
                
                if number_value is not None:
                    sum += number_value
        cell_F1 = sheet['F1']
        value_F1 = convert_string_to_number(str(cell_F1.value))
        if sum == value_F1:
            None
        else:
            print(f"Sheet {sheet_counter + 1} có tổng thành tiền khác với cộng tiền hàng chưa thuế.")
        sheet_counter += 1
    # Lưu lại tệp Excel đã cập nhật
    output_filename_with_time = f'Total_Invoice_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    output_excel_file_with_time = os.path.join("H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_Excel", output_filename_with_time)
    excel.save(output_excel_file_with_time)
    total_invoice_path = os.path.join("H:/APP UNIVERSITY/CODE PYTHON/OpenCV-Master-Computer-Vision-in-Python/SourcecodeOCR/Data_Excel", "Total_Invoice.xlsx")
    os.remove(total_invoice_path)

    excel = openpyxl.load_workbook(output_excel_file_with_time)
    new_sheet_name = "Data_Total"
    new_sheet = excel.create_sheet(new_sheet_name)
    new_sheet['A9'] = "Người nộp thuế: CÔNG TY TNHH EAST - WEST RESTAURANT CONCEPTS"
    new_sheet['A10'] = "Mã số thuế: 0071052970"
    new_sheet['A12'] = "STT"
    new_sheet['B12'] = "Ký hiệu hóa đơn"
    new_sheet['C12'] = "Số hóa đơn"
    new_sheet['D12'] = "Ngày, tháng, năm lập hóa đơn"
    new_sheet['E12'] = "Tên người bán"
    new_sheet['F12'] = "Mã số thuế người bán"
    new_sheet['G12'] = "Mặt hàng"
    new_sheet['H12'] = "Doanh số mua chưa có thuế"
    new_sheet['I12'] = "Thuế GTGT đủ điều kiện khấu trừ thuế"
    new_sheet['J12'] = "Ghi chú"
    for column in new_sheet.iter_cols(min_col=1, max_col=10, min_row=12, max_row=12):
        for cell in column:
            cell.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')
            cell.font = Font(name='Times New Roman', size=12,bold=True)
    for column in new_sheet.iter_cols(min_col=1, max_col=1, min_row=9, max_row=10):
        for cell in column:
            cell.font = Font(name='Times New Roman', size=12, bold=True)
    column_widths = {'A': 5,'B': 10,'C': 10,'D': 12,'E': 85,'F': 20,'G': 40,'H': 15,'I': 15, 'J': 50}
    for column, width in column_widths.items():
        new_sheet.column_dimensions[column].width = width

    excel.save(output_excel_file_with_time)
    count_sheet = 0
    sheets_with_max_row_10 = []
    for sheet_name in excel.sheetnames[:-1]:  
        sheet = excel[sheet_name]
        new_row = len(new_sheet['A'])+1  # Bắt đầu từ hàng 13
        data_to_copy = sheet['D2'].value
        new_sheet.cell(row=new_row, column=2, value=data_to_copy)
        data_to_copy = sheet['D3'].value
        new_sheet.cell(row=new_row, column=3, value=data_to_copy)
        data_to_copy = sheet['D4'].value
        new_sheet.cell(row=new_row, column=4, value=data_to_copy)
        data_to_copy = sheet['B1'].value
        new_sheet.cell(row=new_row, column=5, value=data_to_copy)
        data_to_copy = sheet['B2'].value
        new_sheet.cell(row=new_row, column=6, value=data_to_copy)
        data_to_copy = sheet['F1'].value
        new_sheet.cell(row=new_row, column=8, value=data_to_copy)
        data_to_copy = sheet['F2'].value
        new_sheet.cell(row=new_row, column=9, value=data_to_copy)
        count_sheet +=1
        if sheet.max_row == 10:
            sheets_with_max_row_10.append(sheet_name)
            count_note = count_sheet
            if sheets_with_max_row_10:
                target_sheet_name = sheets_with_max_row_10[0] 
                target_worksheet = excel[target_sheet_name]
                value_at_B10 = target_worksheet['B10'].value
                last_worksheet_name = excel.sheetnames[-1]
                last_worksheet = excel[last_worksheet_name]
                last_worksheet[f'G{12 + count_note}'] = value_at_B10
                excel.save(output_excel_file_with_time)
                sheets_with_max_row_10.pop(0)
        if sheet[f"A{sheet.max_row}"].value == None:
            data_to_copy = sheet[f"B{sheet.max_row}"].value
            new_sheet.cell(row=new_row, column=10, value=data_to_copy)

    for row in range(13, new_sheet.max_row + 1):
        cell_value = new_sheet.cell(row=row, column=7).value
        if ((new_sheet.cell(row=row, column=5).value == "CÔNG TY TRÁCH NHIỆM HỮU HẠN NƯỚC GIẢI KHÁT COCA-COLA VIỆT NAM") and (cell_value == None)) or ((new_sheet.cell(row=row, column=5).value == "CHI NHÁNH CÔNG TY TNHH NƯỚC GIẢI KHÁT COCA-COLA VIỆT NAM TẠI THÀNH PHỐ ĐÀ NẴNG") and (cell_value == None)):
            new_sheet.cell(row=row, column=7, value="Chi phí mua đồ uống giải khát")
        elif new_sheet.cell(row=row, column=5).value == "CÔNG TY TNHH THỰC PHẨM LINH KHOA" and cell_value == None:
            new_sheet.cell(row=row, column=7, value="Chi phí mua thực phẩm")
        elif new_sheet.cell(row=row, column=5).value == "CN CÔNG TY TNHH NƯỚC GIẢI KHÁT SUNTORY PEPSICO VIỆT NAM TẠI TỈNH BẮC NINH" and cell_value == None:
            new_sheet.cell(row=row, column=7, value="Chi phí mua đồ uống giải khát")    

    for column in new_sheet.iter_cols(min_col=4, max_col=4, min_row=13, max_row=new_sheet.max_row):
        for cell in column:
            cell.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')
    for column in new_sheet.iter_cols(min_col=1, max_col=1, min_row=13, max_row=new_sheet.max_row):
        for cell in column:
            cell.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')
    for index, number in enumerate(range(1, count_sheet + 1)):
        destination_row = 13 + index
        new_sheet.cell(row=destination_row, column=1, value=number)
   
    number_style = NamedStyle(name='number_style', number_format='#,##0')
    for row in new_sheet.iter_rows(min_row=13, max_row=new_sheet.max_row, min_col=8, max_col=9):
        for cell in row:
            cell.style = number_style
    for column in new_sheet.iter_cols(min_col=8, max_col=9, min_row=13, max_row=new_sheet.max_row):
        for cell in column:
            cell.alignment = Alignment(wrapText=True,horizontal='right', vertical='center')
    for column in new_sheet.iter_cols(min_col=1, max_col=10, min_row=13, max_row=new_sheet.max_row):
        for cell in column:
            cell.font = Font(name='Times New Roman', size=10)

    output_folder = "H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Data_Final_Excel"
    output_excel_file_with_time1 = os.path.join(output_folder, output_filename_with_time)
    excel.save(output_excel_file_with_time1)
    workbook = load_workbook(output_excel_file_with_time1)
    worksheet_names = workbook.sheetnames
    new_names = [f"Invoice{i+1}" for i in range(len(worksheet_names)-1)]
    for i in range(len(worksheet_names)-1):
        worksheet = workbook[worksheet_names[i]]
        worksheet.title = new_names[i]
    workbook.save(output_excel_file_with_time1)
    os.remove(output_excel_file_with_time)
    folder_to_clear = 'SourcecodeOCR/Data_Excel'
    for filename in os.listdir(folder_to_clear):
        file_path = os.path.join(folder_to_clear, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)
    os.startfile(output_folder)

def pdf_to_png_INV():
    poppler_path = r'H:/OCR/Popler/poppler-23.07.0/Library/bin'
    pdf_folder = r"H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\File_PDF_Invoice"
    saving_folder = r"H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Compare_Invoice"
    os.makedirs(saving_folder, exist_ok=True)
    for pdf_filename in os.listdir(pdf_folder):
        if pdf_filename.lower().endswith(".pdf"):
            pdf_path = os.path.join(pdf_folder, pdf_filename)
            pages = convert_from_path(pdf_path=pdf_path, poppler_path=poppler_path)
            for c, page in enumerate(pages, start=1):
                img_name = f"{os.path.splitext(pdf_filename)[0]}_Page{c}.png"
                img_path = os.path.join(saving_folder, img_name)
                page.save(img_path, "png")
        os.remove(pdf_path)

#Xóa file ảnh
def delete_img_files(folder_path):
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except Exception as e:
                None

#Xóa file pdf
def delete_pdf_files(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith(".pdf"):
            file_path = os.path.join(folder_path, filename)
            os.remove(file_path)

#Xác thực nhà cung cấp
def verify_invoice(mst1,ms1,kh1,shd1):  
    chrome_options = webdriver.ChromeOptions()  # Đổi tên biến thành chrome_options
    chrome_options.add_argument('--headless')
    web = webdriver.Chrome(options=chrome_options)  # Chỉ sử dụng options=chrome_options
    web.get("https://tracuuhoadon.gdt.gov.vn/search1hd.html")
    sleep(1)
    mst = web.find_element(By.ID,"tin")
    ms = web.find_element(By.ID,"mau")
    kh = web.find_element(By.ID,"kyhieu")
    shd = web.find_element(By.ID,"so")
    btn = web.find_element(By.ID,"searchBtn")
    capcha = web.find_element(By.ID,"captchaCodeVerify")
    sleep(1)
    mst.send_keys(mst1)   
    ms.send_keys(ms1)
    kh.send_keys(kh1)
    shd.send_keys(shd1)
    sleep(1)
    web.get_screenshot_as_file("capcha.png")
    sleep(1)
    img_web = cv2.imread("capcha.png")
    imgcrop = img_web[386:413,378:500]
    #imgcrop = img_web[470:505,504:658]
    cv2.imwrite("Captcha123.png",imgcrop)
    new_image_path = "H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\Captcha123.png"
    predicted_text = predict_captcha(new_image_path)
    capcha.send_keys(predicted_text)
    btn.click()
    sleep(1)
    mXt = web.find_element(By.ID,"lbLoiCode")
    if mXt == "Sai mã xác thực!":
        for i in range(1,10):
            web.get("https://tracuuhoadon.gdt.gov.vn/search1hd.html")
            sleep(1)
            mst = web.find_element(By.ID,"tin")
            ms = web.find_element(By.ID,"mau")
            kh = web.find_element(By.ID,"kyhieu")
            shd = web.find_element(By.ID,"so")
            btn = web.find_element(By.ID,"searchBtn")
            capcha = web.find_element(By.ID,"captchaCodeVerify")
            sleep(1)
            mst.send_keys(mst1)   
            ms.send_keys(ms1)
            kh.send_keys(kh1)
            shd.send_keys(shd1)
            sleep(1)
            web.get_screenshot_as_file("capcha.png")
            sleep(1)
            img_web = cv2.imread("capcha.png")
            imgcrop = img_web[386:413,378:500]
            cv2.imwrite("Captcha123.png",imgcrop)
            new_image_path = "H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\Captcha123.png"
            predicted_text = predict_captcha(new_image_path)
            capcha.send_keys(predicted_text)
            btn.click()
            sleep(1)
            mXt = web.find_element(By.ID,"lbLoiCode")
            if mXt != "Sai mã xác thực!":
                web.get_screenshot_as_file("new_page_screenshot.png")
                sleep(1)
                img_web1 = cv2.imread("new_page_screenshot.png")
                imgcrop1 = img_web1[314:337,293:685]
                cv2.imwrite("Xacthuc.png",imgcrop1)
                text1 = pytesseract.image_to_string(imgcrop1, lang='vie', config='--oem 3 --psm 6')
                text1 = text1.strip().replace('\n', '').replace('.','')
                return text1
    else:        
        web.get_screenshot_as_file("new_page_screenshot.png")
        sleep(1)
        img_web1 = cv2.imread("new_page_screenshot.png")
        imgcrop1 = img_web1[314:337,293:685]
        cv2.imwrite("Xacthuc.png",imgcrop1)
        text1 = pytesseract.image_to_string(imgcrop1, lang='vie', config='--oem 3 --psm 6')
        text1 = text1.strip().replace('\n', '')
        return text1

#Chuyển đổi ảnh sang pdf
def convert_image_to_pdf(image_path):
    pdf_folder = r"SourcecodeOCR\File_PDF_Invoice"
    image_name = os.path.basename(image_path)
    pdf_name = os.path.splitext(image_name)[0] + ".pdf"
    pdf_path = os.path.join(pdf_folder, pdf_name)
    with open(pdf_path, "wb") as pdf_file:
        pdf_file.write(img2pdf.convert(image_path))

#Kiểm tra định dạng ngày/tháng/năm
def is_date_format(string):
    try:
        datetime.strptime(string, '%d/%m/%Y')
        return True
    except ValueError:
        return False

#Định dạng data đơn giá
def format_string_dg(number_str):
    clean_str = ''.join(filter(lambda char: char.isdigit() or char == ',', number_str))
    number = int(clean_str.replace(',', ''))
    formatted_str = '{:,.2f}'.format(number / 100).replace('.', ',')
    return formatted_str

#Định dạng data thành tiền
def format_string_tt(number_str):
    first_dot_index = number_str.find('.')
    second_dot_index = number_str.find('.', first_dot_index + 1)
    if second_dot_index != -1:
        number_str = number_str[:second_dot_index] + ',' + number_str[second_dot_index + 1:]
    return number_str

def invoice_cocacola(img,supplier,mst1,address1,consumer,mst2,address2,ms,kh,so,ngaygiao):
    wb = xw.Book()
    sht = wb.sheets.active

    sht.range("A1").value = 'Nhà cung cấp'
    sht.range("B1").value = supplier

    sht.range("A2").value = 'Mã số thuế'
    sht.range("B2").number_format = "@"
    sht.range("B2").value = mst1

    sht.range("A3").value = 'Địa chỉ'
    sht.range("B3").value = address1

    sht.range("A5").value = 'Người mua hàng'
    sht.range("B5").value = consumer

    sht.range("A6").value = 'Mã số thuế'
    sht.range("B6").number_format = "@"
    sht.range("B6").value = mst2

    sht.range("A7").value = 'Địa chỉ'
    sht.range("B7").value = address2

    sht.range("C1").value = 'Mẫu số'
    sht.range("D1").number_format = "@"
    sht.range("D1").value = ms

    sht.range("C2").value = 'Ký hiệu'
    sht.range("D2").number_format = "@"
    sht.range("D2").value = kh

    sht.range("C3").value = 'Số'
    sht.range("D3").number_format = "@"
    sht.range("D3").value = so

    sht.range("C4").value = 'Ngày giao'
    sht.range("D4").number_format = "@"
    sht.range("D4").value = ngaygiao


    #STT
    list_stt = []
    img_stt = img[1142:1620, 102:163]
    data_stt = pytesseract.image_to_string(img_stt, lang = 'vie',config= '--oem 3 --psm 6')
    lines_data_stt = data_stt.strip().splitlines()
    pattern = re.compile(r'^\d+$')
    numbers_stt = [element for element in lines_data_stt if pattern.search(element)]
    for i in numbers_stt:
        list_stt.append(i)
    count = len(list_stt)
    sht.range("A9").value = 'STT'
    for index, line in enumerate(list_stt, start=10):
        sht[f"A{index}"].number_format = "@"  
        sht[f"A{index}"].value = line 

    list_hh = []
    count_hh1 = count
    #Tên hàng hóa, dịch vụ
    img_hh = img[1138:1930, 163:836]
    data_hh = pytesseract.image_to_string(img_hh, lang = 'vie',config= '--oem 3 --psm 6')
    lines_data_hh = data_hh.strip().splitlines()
    elements_data_hh = [element[:-1] if element.endswith('.') or element.endswith(',') else element for element in lines_data_hh]
    elements_data_hh = [element.replace("mẾng", "uống").replace("…Ếng","uống") for element in elements_data_hh]
    sht.range("B9").value = "Tên hàng hóa, dịch vụ"
    for index, i in enumerate(elements_data_hh):
        if index <= count_hh1 -1:
            list_hh.append(i)
        if index == count_hh1:
            if "thay thế" in elements_data_hh[index]:
                list_hh.append(elements_data_hh[index])
    for index, line in enumerate(list_hh, start=10):  
        sht[f"B{index}"].value = line 

    #Đơn vị tính
    img_dvt = img[1140:1930, 965:1095]
    data_dvt = pytesseract.image_to_string(img_dvt, lang = 'vie',config= '--oem 3 --psm 6')
    lines_data_dvt = data_dvt.strip().splitlines()
    filtered_lines = [line for line in lines_data_dvt if line.strip() != "" and line.strip() != " "]
    sht.range("C9").value = 'Đơn vị tính'
    for index, line in enumerate(filtered_lines[:count], start=10):
        sht[f"C{index}"].value = line

    #Số lượng
    img_sl = img[1138:1930, 1180:1220]
    text = pytesseract.image_to_string(img_sl, lang = 'vie',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    pattern = re.compile(r'\d+')
    elements_with_numbers = [element for element in lines if pattern.search(element)]
    sht.range("D9").value = 'Số lượng'
    for index, line in enumerate(elements_with_numbers[:count], start=10):
        sht[f"D{index}"].number_format = "@"
        sht[f"D{index}"].value = line

    #Đơn giá
    img_dg = img[1138:1930, 1221:1340]
    text = pytesseract.image_to_string(img_dg, lang = 'vie',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    sht.range("E9").value = 'Đơn giá'
    for index, line in enumerate(lines[:count], start=10):
        sht[f"E{index}"].number_format = "#,##0"
        sht[f"E{index}"].value = line

    #Thành tiền
    #half_image2 = img[1140:1929, 1400:1553]
    img_tt = img[1140:1929, 1400:1555]
    img_tt = cv2.resize(img_tt,None,fx=1.4,fy=1.4,interpolation=cv2.INTER_LINEAR)
    text = pytesseract.image_to_string(img_tt, lang = 'vie',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    sht.range("F9").value = 'Thành tiền'
    for index, line in enumerate(lines[:count], start=10):
        sht[f"F{index}"].number_format = "#,##0"
        line = line.replace('.', ',') if '.' in line else line
        line = float(line.replace(',', '')) if ',' in line else float(line)
        sht[f"F{index}"].value = line
        if line == 0:
            sht[f"E{index}"].value = line

    #Total
    img_total = img[1335:2120, 630:1555]
    img_total = cv2.resize(img_total,None,fx=1.5,fy=1.5,interpolation=cv2.INTER_BITS)
    img_total = cv2.cvtColor(img_total,cv2.COLOR_BGR2GRAY)
    kernel = np.array([[-1, -1, -1],
                        [-1,  9, -1],
                        [-1, -1, -1]])
    img_total = cv2.filter2D(img_total, -1, kernel)
    result_2_1 = pytesseract.image_to_data(img_total, lang = 'vie+eng', config= 'tessdata',output_type= Output.DICT)
    
    filtered_data = [item for item in result_2_1['text'] if item.strip() != '' and item.strip() != ' ']
    numeric_elements = [element for element in filtered_data if any(char.isdigit() or char == '\\' for char in element) and '(' not in element and ')' not in element]

    value = numeric_elements[0].replace('.', ',') if '.' in numeric_elements[0] else numeric_elements[0]
    sht.range("E1").value = 'Cộng tiền hàng chưa có thuế GTGT'
    sht.range("F1").number_format = "#,##0"
    text_value = value
    numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
    sht.range("F1").value = numeric_value
    
    value = numeric_elements[1].replace('.', ',') if '.' in numeric_elements[1] else numeric_elements[1]
    sht.range("E2").value = 'Tiền thuế GTGT'
    sht.range("F2").number_format = "#,##0"
    if value.replace(',', '').replace('.', '').isdigit():
        text_value = value
        numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
    else:
        # Nếu value không phải là số, giữ nguyên value
        text_value = value
        numeric_value = text_value
    sht.range("F2").value = numeric_value

    value = numeric_elements[2].replace('.', ',') if '.' in numeric_elements[2] else numeric_elements[2]
    sht.range("E3").value = 'Tổng cộng tiền thanh toán'
    sht.range("F3").number_format = "#,##0"
    text_value = value
    numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
    sht.range("F3").value = numeric_value


    path_compare1 = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Compare_Invoice'
    path_source1 = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Images_Invoice'
    orb = cv2.ORB_create(nfeatures = 1000)
    images_compare1 = []
    images_source1 = []
    myList_compare1 = os.listdir(path_compare1)
    mylist_source1 = os.listdir(path_source1)
    for img1 in myList_compare1:
        images_compare1 = cv2.imread(f'{path_compare1}/{img1}')
        images_compare1 = cv2.cvtColor(images_compare1,cv2.COLOR_BGR2GRAY)
        kp1, des1 = orb.detectAndCompute(images_compare1,None)
        matchList1 = []
        count1 = count
        for img2 in mylist_source1:
            index_1 = myList_compare1.index(img1)
            images_source1 = cv2.imread(f'{path_source1}/{img2}')
            kp2, des2 = orb.detectAndCompute(images_source1,None)
            bf = cv2.BFMatcher()
            matches = bf.knnMatch(des1,des2,k=2)
            good = []
            for m,n in matches:
                if m.distance < 0.75*n.distance:
                    good.append([m])
            matchList1.append(len(good))
        if max(matchList1) > 450:
            img_path = os.path.join(path_compare1, myList_compare1[index_1])
            list_stt1 = []
            img_stt = img[1142:1620, 102:163]
            text = pytesseract.image_to_string(img_stt, lang = 'vie',config= '--oem 3 --psm 6')
            lines = text.strip().splitlines()
            pattern = re.compile(r'^\d+$')
            elements_with_numbers = [element for element in lines if pattern.search(element)]
            for i in elements_with_numbers:
                list_stt1.append(i)
            count = len(list_stt1)
            if count == 0:
                img_total = img[1335:2120, 630:1555]
                img_total = cv2.resize(img_total,None,fx=1.5,fy=1.5,interpolation=cv2.INTER_BITS)
                img_total = cv2.cvtColor(img_total,cv2.COLOR_BGR2GRAY)
                kernel = np.array([[-1, -1, -1],
                                    [-1,  9, -1],
                                    [-1, -1, -1]])
                img_total = cv2.filter2D(img_total, -1, kernel)
                result_2_1 = pytesseract.image_to_data(img_total, lang = 'vie+eng', config= 'tessdata',output_type= Output.DICT)
                #result_2_1 = pytesseract.image_to_data(half_image2, lang = 'vie', config= 'tessdata',output_type= Output.DICT)
                filtered_data = [item for item in result_2_1['text'] if item.strip() != '' and item.strip() != ' ']
                numeric_elements = [element for element in filtered_data if any(char.isdigit() or char == '\\' for char in element) and '(' not in element and ')' not in element]
                value = numeric_elements[0].replace('.', ',') if '.' in numeric_elements[0] else numeric_elements[0]
                sht.range("E1").value = 'Cộng tiền hàng chưa có thuế GTGT'
                sht.range("F1").number_format = "#,##0"
                text_value = value
                numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
                sht.range("F1").value = numeric_value
                
                value = numeric_elements[1].replace('.', ',') if '.' in numeric_elements[1] else numeric_elements[1]
                sht.range("E2").value = 'Tiền thuế GTGT'
                sht.range("F2").number_format = "#,##0"
                if value.replace(',', '').replace('.', '').isdigit():
                    text_value = value
                    numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
                else:
                    # Nếu value không phải là số, giữ nguyên value
                    text_value = value
                    numeric_value = text_value
                sht.range("F2").value = numeric_value

                value = numeric_elements[2].replace('.', ',') if '.' in numeric_elements[2] else numeric_elements[2]
                sht.range("E3").value = 'Tổng cộng tiền thanh toán'
                sht.range("F3").number_format = "#,##0"
                text_value = value
                numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
                sht.range("F3").value = numeric_value

            else:
                sht.range("A9").value = 'STT'
                for index, line in enumerate(list_stt1, start=10 + count1):
                    sht[f"A{index}"].number_format = "@"  
                    sht[f"A{index}"].value = line 

                list_hh1 = []
                count_hh1 = count
                #Tên hàng hóa, dịch vụ
                img_hh = img[1138:1930, 163:836]
                text = pytesseract.image_to_string(img_hh, lang = 'vie',config= '--oem 3 --psm 6')
                lines = text.strip().splitlines()
                drink_elements = [element[:-1] if element.endswith('.') or element.endswith(',') else element for element in lines]
                drink_elements = drink_elements.replace("mẾng","uống")
                sht.range("B9").value = "Tên hàng hóa, dịch vụ"
                for index, i in enumerate(drink_elements):
                    if index <= count_hh1 -1:
                        list_hh.append(i)
                    if index == count_hh1:
                        if "thay thế" in drink_elements[index]:
                            list_hh.append(drink_elements[index])
                for index, line in enumerate(list_hh1, start=10 + count1):  
                    sht[f"B{index}"].value = line 
                #Đơn vị tính
                img_dvt = img[1138:1930, 965:1095]
                text = pytesseract.image_to_string(img_dvt, lang = 'vie',config= '--oem 3 --psm 6')
                lines = text.strip().splitlines()
                filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
                sht.range("C9").value = 'Đơn vị tính'
                for index, line in enumerate(filtered_lines[:count], start=10 + count1):
                    sht[f"C{index}"].value = line
                #Số lượng
                img_sl = img[1138:1930, 1180:1220]
                text = pytesseract.image_to_string(img_sl, lang = 'vie',config= '--oem 3 --psm 6')
                lines = text.strip().splitlines()
                pattern = re.compile(r'\d+')
                elements_with_numbers = [element for element in lines if pattern.search(element)]
                sht.range("D9").value = 'Số lượng'
                for index, line in enumerate(elements_with_numbers[:count], start=10 + count1):
                    sht[f"D{index}"].number_format = "@"
                    sht[f"D{index}"].value = line
                #Đơn giá 
                img_dg = img[1138:1930, 1221:1340]
                text = pytesseract.image_to_string(img_dg, lang = 'vie',config= '--oem 3 --psm 6')
                lines = text.strip().splitlines()
                sht.range("E9").value = 'Đơn giá'
                for index, line in enumerate(lines[:count], start=10 + count1):
                    sht[f"E{index}"].number_format = "@"
                    sht[f"E{index}"].value = line
                #Thành tiền
                #half_image2 = img[1140:1929, 1400:1553]
                img_tt = img[1140:1929, 1400:1555]
                img_tt = cv2.resize(img_tt,None,fx=1.4,fy=1.4,interpolation=cv2.INTER_LINEAR)
                text = pytesseract.image_to_string(img_tt, lang = 'vie',config= '--oem 3 --psm 6')
                lines = text.strip().splitlines()
                sht.range("F9").value = 'Thành tiền'
                for index, line in enumerate(lines[:count], start=10):
                    sht[f"F{index}"].number_format = "@"
                    line = line.replace('.', ',') if '.' in line else line
                    line = float(line.replace(',', '')) if ',' in line else float(line)
                    sht[f"F{index}"].value = line
                    if line == 0:
                        sht[f"E{index}"].value = line

                img_total = img[1335:2120, 630:1555]
                img_total = cv2.resize(img_total,None,fx=1.5,fy=1.5,interpolation=cv2.INTER_BITS)
                img_total = cv2.cvtColor(img_total,cv2.COLOR_BGR2GRAY)
                kernel = np.array([[-1, -1, -1],
                                    [-1,  9, -1],
                                    [-1, -1, -1]])
                img_total = cv2.filter2D(img_total, -1, kernel)
                result_2_1 = pytesseract.image_to_data(img_total, lang = 'vie+eng', config= 'tessdata',output_type= Output.DICT)
                filtered_data = [item for item in result_2_1['text'] if item.strip() != '' and item.strip() != ' ']
                numeric_elements = [element for element in filtered_data if any(char.isdigit() or char == '\\' for char in element) and '(' not in element and ')' not in element]
                value = numeric_elements[0].replace('.', ',') if '.' in numeric_elements[0] else numeric_elements[0]
                sht.range("E1").value = 'Cộng tiền hàng chưa có thuế GTGT'
                sht.range("F1").number_format = "#,##0"
                text_value = value
                numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
                sht.range("F1").value = numeric_value
                
                value = numeric_elements[1].replace('.', ',') if '.' in numeric_elements[1] else numeric_elements[1]
                sht.range("E2").value = 'Tiền thuế GTGT'
                sht.range("F2").number_format = "#,##0"
                if value.replace(',', '').replace('.', '').isdigit():
                    text_value = value
                    numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
                else:
                    text_value = value
                    numeric_value = text_value
                sht.range("F2").value = numeric_value

                value = numeric_elements[2].replace('.', ',') if '.' in numeric_elements[2] else numeric_elements[2]
                sht.range("E3").value = 'Tổng cộng tiền thanh toán'
                sht.range("F3").number_format = "#,##0"
                text_value = value
                numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
                sht.range("F3").value = numeric_value
            os.remove(img_path)
        else:
            img_path = os.path.join(path_compare1, myList_compare1[index_1])
            os.remove(img_path)
            break
    file_name = f"{so}_Coca_INV.xlsx"
    wb.save(fr'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Data_Excel\{file_name}')
    excel_file_path = fr'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Data_Excel\{file_name}'  
    columns_bold(excel_file_path, 1,7,1)
    columns_bold(excel_file_path, 1,5,3)
    columns_bold(excel_file_path, 1,4,5)
    rows_bold(excel_file_path, 1, 6, 9)
    wb = xw.Book(excel_file_path)
    change_font_size(wb, 12)
    process_all_worksheets(excel_file_path)

def invoice_linhkhoa(img,supplier,mst1,address1,consumer,mst2,address2,ms,kh,so,ngaygiao):  
    wb = xw.Book()
    sht = wb.sheets.active

    sht.range("A1").value = 'Nhà cung cấp'
    sht.range("B1").value = supplier

    sht.range("A2").value = 'Mã số thuế'
    sht.range("B2").number_format = "@"
    sht.range("B2").value = mst1

    sht.range("A3").value = 'Địa chỉ'
    sht.range("B3").value = address1

    sht.range("A5").value = 'Người mua hàng'
    sht.range("B5").value = consumer

    sht.range("A6").value = 'Mã số thuế'
    sht.range("B6").number_format = "@"
    sht.range("B6").value = mst2

    sht.range("A7").value = 'Địa chỉ'
    sht.range("B7").value = address2

    sht.range("C1").value = 'Mẫu số'
    sht.range("D1").number_format = "@"
    sht.range("D1").value = ms

    sht.range("C2").value = 'Ký hiệu'
    sht.range("D2").number_format = "@"
    sht.range("D2").value = kh

    sht.range("C3").value = 'Số'
    sht.range("D3").number_format = "@"
    sht.range("D3").value = so

    sht.range("C4").value = 'Ngày giao'
    sht.range("D4").number_format = "@"
    sht.range("D4").value = ngaygiao

    
    list_stt = []
    img1 = img[897:1620, 90:175]
    text = pytesseract.image_to_string(img1, lang = 'eng',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    pattern = re.compile(r'^\d+$')
    elements_with_numbers = [element for element in lines if pattern.search(element)]
    for i in elements_with_numbers:
        list_stt.append(i)
    img2 = img[1620:2200, 90:175]
    text = pytesseract.image_to_string(img2, lang = 'eng',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    elements_with_numbers = [element for element in lines if pattern.search(element)]
    for i in elements_with_numbers:
        list_stt.append(i)
    count = len(list_stt)
    sht.range("A9").value = 'STT'
    for index, line in enumerate(list_stt, start=10):
        sht[f"A{index}"].number_format = "@"  
        sht[f"A{index}"].value = line 

    list_hh = []
    count_hh1 = count
    img1 = img[897:1620, 175:776]
    img1 = cv2.cvtColor(img1, cv2.COLOR_RGB2GRAY)
    img2 = img1[418:512,336:597]
    _, thresholded_roi1 = cv2.threshold(img2, 150, 255, cv2.THRESH_BINARY) #150
    img2 = img1[237:384,426:598]
    _, thresholded_roi2 = cv2.threshold(img2, 150, 255, cv2.THRESH_BINARY) #150
    img2 = img1[78:225,565:598]
    _, thresholded_roi3 = cv2.threshold(img2, 150, 255, cv2.THRESH_BINARY) #150
    img1[418:512,336:597] = thresholded_roi1
    img1[237:384,426:598] = thresholded_roi2
    img1[78:225,565:598] = thresholded_roi3
    text = pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    filtered_data = [item for item in lines if item.strip() != '']
    filtered_data = [s.replace("NÁM", "NẤM").replace("Rết", "Rốt").replace("Is", "1s") for s in filtered_data]
    for i in filtered_data[:count_hh1]:
        list_hh.append(i)
        count_hh1 -= 1
    img2 = img[1620:2200, 175:776]
    text = pytesseract.image_to_string(img2, lang = 'vie',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    for i in lines[:count_hh1]:
        list_hh.append(i)
        count_hh1 -= 1
    sht.range("B9").value = 'Tên hàng hóa, dịch vụ'
    for index, line in enumerate(list_hh, start=10):  
        sht[f"B{index}"].value = line 

    list_dvt = []
    count_dvt1 = count
    img1 = img[897:1620, 777:832]
    img1 = cv2.cvtColor(img1, cv2.COLOR_RGB2GRAY)
    text = pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
    for i in filtered_lines[:count_dvt1]:
        if i.startswith('k'):
            i = 'kg'
            list_dvt.append(i)
            count_dvt1 -= 1
        elif i.startswith('Qủa'):
            list_dvt.append(i)
            count_dvt1 -= 1
    img2 = img[897:1620, 777:832]
    img2 = cv2.cvtColor(img2, cv2.COLOR_RGB2GRAY)
    text = pytesseract.image_to_string(img2, lang = 'vie',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
    for i in filtered_lines[:count_dvt1]:
        if i.startswith('k'):
            i = 'kg'
            list_dvt.append(i)
            count_dvt1 -= 1
        elif i.startswith('Qủa'):
            list_dvt.append(i)
            count_dvt1 -= 1
    sht.range("C9").value = 'Đơn vị tính'
    for index, line in enumerate(list_dvt, start=10):  
        sht[f"C{index}"].value = line 
    
    list_sl = []
    count_sl1 = count
    img1 = img[897:1620, 1000:1115]
    img1 = cv2.cvtColor(img1, cv2.COLOR_RGB2GRAY)
    text = pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
    for i in filtered_lines[:count_sl1]:
        i = format_string_dg(i)
        list_sl.append(i)
        count_sl1 -= 1
    img2 = img[1620:2200, 1000:1115]
    text = pytesseract.image_to_string(img2, lang = 'vie',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
    for i in filtered_lines[:count_sl1]:
        i = format_string_dg(i)
        list_sl.append(i)
        count_sl1 -= 1
    sht.range("D9").value = 'Số lượng'
    for index, line in enumerate(list_sl, start=10):  
        sht[f"D{index}"].number_format = "@"
        sht[f"D{index}"].value = line 

    list_dg = []
    count_dg1 = count
    img1 = img[897:1620, 1135:1327]
    text = pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
    for i in filtered_lines[:count_dg1]:
        i = format_string_tt(i)
        list_dg.append(i)
        count_dg1 -= 1
    img2 = img[1620:2200, 1135:1327]
    text = pytesseract.image_to_string(img2, lang = 'vie',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
    for i in filtered_lines[:count_dg1]:
        i = format_string_tt(i)
        list_dg.append(i)
        count_dg1 -= 1
    sht.range("E9").value = 'Đơn giá'
    for index, line in enumerate(list_dg, start=10):  
        sht[f"E{index}"].number_format = "@"
        sht[f"E{index}"].value = line

    list_tt = []
    count_tt1 = count
    img1 = img[897:1620, 1400:1568]
    text = pytesseract.image_to_string(img1, lang = 'eng',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
    for i in filtered_lines[:count_tt1]:
        list_tt.append(i)
        count_tt1 -= 1
    img2 = img[1620:2200, 1400:1568]
    text = pytesseract.image_to_string(img2, lang = 'eng',config= '--oem 3 --psm 6')
    lines = text.strip().splitlines()
    filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
    for i in filtered_lines[:count_tt1]:
        list_tt.append(i)
        count_tt1 -= 1
    sht.range("F9").value = 'Thành tiền'
    for index, line in enumerate(list_tt, start=10):  
        sht[f"F{index}"].number_format = "#,##0"
        text_value = line
        numeric_value = float(text_value.replace('.', '').replace(',', ''))
        sht[f"F{index}"].value = numeric_value

    img1 = img[1470:2200, 825:1570]
    result_2_1 = pytesseract.image_to_data(img1, lang = 'vie', config= 'tessdata',output_type= Output.DICT)
    filtered_data = [item for item in result_2_1['text'] if item.strip() != '' and item.strip() != ' ']
    for i, item in enumerate(filtered_data):
        if item == "Cộng":
            sht.range("E1").value = 'Cộng tiền hàng chưa có thuế GTGT'
            sht.range("F1").number_format = "#,##0"
            value = filtered_data[i + 3].replace(',', '.')
            numeric_value = float(value.replace('.', '').replace(',', ''))
            sht.range("F1").value = numeric_value
        elif item == "Tiền":
            sht.range("E2").value = 'Tiền thuế GTGT'
            sht.range("F2").number_format = "#,##0"
            value = filtered_data[i + 3].replace(',', '.')
            if value == "X" or value == "x":
                sht.range("F2").value = value
            else:
                numeric_value = float(value.replace('.', '').replace(',', ''))
                sht.range("F2").value = numeric_value
        elif item == "thanh":
            sht.range("E3").value = 'Tổng cộng tiền thanh toán'
            sht.range("F3").number_format = "#,##0"
            value = filtered_data[i + 2].replace(',', '.')
            numeric_value = float(value.replace('.', '').replace(',', ''))
            sht.range("F3").value = numeric_value
    
    path_compare1 = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Compare_Invoice'
    path_source1 = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Images_Invoice'
    orb = cv2.ORB_create(nfeatures = 1000)
    images_compare1 = []
    images_source1 = []
    myList_compare1 = os.listdir(path_compare1)
    mylist_source1 = os.listdir(path_source1)
    for index in range(len(myList_compare1)):
        if len(myList_compare1) > 0:
            img1 = myList_compare1[0]
            if "Page1" in img1:
                break
            else:
                images_compare1 = cv2.imread(f'{path_compare1}/{img1}')
                images_compare1 = cv2.cvtColor(images_compare1,cv2.COLOR_BGR2GRAY)
                kp1, des1 = orb.detectAndCompute(images_compare1,None)
                matchList1 = []
                count1 = count
                for img2 in mylist_source1:
                    index_1 = myList_compare1.index(img1)
                    images_source1 = cv2.imread(f'{path_source1}/{img2}')
                    kp2, des2 = orb.detectAndCompute(images_source1,None)
                    bf = cv2.BFMatcher()
                    matches = bf.knnMatch(des1,des2,k=2)
                    good = []
                    for m,n in matches:
                        if m.distance < 0.75*n.distance:
                            good.append([m])
                    matchList1.append(len(good))
                if max(matchList1) > 350:
                    img_path = os.path.join(path_compare1, myList_compare1[index_1])
                    img = cv2.imread(img_path)
                    img1 = img[897:1620, 88:177]
                    lines = text.strip().splitlines()
                    pattern = re.compile(r'^\d+$')
                    elements_with_numbers = [element for element in lines if pattern.search(element)]
                    if not elements_with_numbers or (not pattern.match(elements_with_numbers[0])):
                        img1 = img[801:1300, 78:1573]
                        result_2_1 = pytesseract.image_to_data(img1, lang = 'vie+eng', config= 'tessdata',output_type= Output.DICT)
                        filtered_data = [item for item in result_2_1['text'] if item.strip() != '' and item.strip() != ' ']
                        for i, item in enumerate(filtered_data):
                            if item == "Cộng":
                                sht.range("E1").value = 'Cộng tiền hàng chưa có thuế GTGT'
                                sht.range("F1").number_format = "#,##0"
                                value = filtered_data[i + 3].replace(',', '.')
                                numeric_value = float(value.replace('.', '').replace(',', ''))
                                sht.range("F1").value = numeric_value
                            elif item == "Tiền":
                                sht.range("E2").value = 'Tiền thuế GTGT'
                                sht.range("F2").number_format = "#,##0"
                                value = filtered_data[i + 3].replace(',', '.')
                                if value == "X" or value == "x":
                                    sht.range("F2").value = value
                                else:
                                    numeric_value = float(value.replace('.', '').replace(',', ''))
                                    sht.range("F2").value = numeric_value
                            elif item == "thanh":
                                sht.range("E3").value = 'Tổng cộng tiền thanh toán'
                                sht.range("F3").number_format = "#,##0"
                                value = filtered_data[i + 2].replace(',', '.')
                                numeric_value = float(value.replace('.', '').replace(',', ''))
                                sht.range("F3").value = numeric_value

                        img_path = os.path.join(path_compare1, myList_compare1[index_1])
                        os.remove(img_path)
                        break
                    else:
                        list_stt = []
                        img1 = img[897:1620, 88:177]
                        text = pytesseract.image_to_string(img1, lang = 'eng',config= '--oem 3 --psm 6')
                        lines = text.strip().splitlines()
                        pattern = re.compile(r'^\d+$')
                        elements_with_numbers = [element for element in lines if pattern.search(element)]
                        for i in elements_with_numbers:
                            list_stt.append(i)
                        img2 = img[1620:2200, 88:177]
                        text = pytesseract.image_to_string(img2, lang = 'eng',config= '--oem 3 --psm 6')
                        lines = text.strip().splitlines()
                        elements_with_numbers = [element for element in lines if pattern.search(element)]
                        for i in elements_with_numbers:
                            list_stt.append(i)
                        count = len(list_stt)
                        sht.range("A9").value = 'STT'
                        for index, line in enumerate(list_stt, start=10 + count1):
                            sht[f"A{index}"].number_format = "@"  
                            sht[f"A{index}"].value = line 

                        list_hh = []
                        count_hh1 = count
                        img1 = img[897:1620, 175:776]
                        img1 = cv2.cvtColor(img1, cv2.COLOR_RGB2GRAY)
                        img2 = img1[418:512,336:597]
                        _, thresholded_roi1 = cv2.threshold(img2, 150, 255, cv2.THRESH_BINARY) #150
                        img2 = img1[237:384,426:598]
                        _, thresholded_roi2 = cv2.threshold(img2, 150, 255, cv2.THRESH_BINARY) #150
                        img2 = img1[78:225,565:598]
                        _, thresholded_roi3 = cv2.threshold(img2, 150, 255, cv2.THRESH_BINARY) #150
                        img1[418:512,336:597] = thresholded_roi1
                        img1[237:384,426:598] = thresholded_roi2
                        img1[78:225,565:598] = thresholded_roi3
                        text = pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6')
                        lines = text.strip().splitlines()
                        filtered_data = [item for item in lines if item.strip() != '']
                        filtered_data = [s.replace("NÁM", "NẤM").replace("Rết", "Rốt").replace("Is", "1s") for s in filtered_data]
                        for i in filtered_data[:count_hh1]:
                            list_hh.append(i)
                            count_hh1 -= 1
                        sht.range("B9").value = 'Tên hàng hóa, dịch vụ'
                        for index, line in enumerate(list_hh, start=10 + count1):  
                            sht[f"B{index}"].value = line 

                        list_dvt = []
                        count_dvt1 = count
                        img1 = img[897:1620, 770:832]
                        img1 = cv2.cvtColor(img1, cv2.COLOR_RGB2GRAY)
                        text = pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6')
                        lines = text.strip().splitlines()
                        filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
                        for i in filtered_lines[:count_dvt1]:
                            if i.startswith('k'):
                                i = 'kg'
                                list_dvt.append(i)
                                count_dvt1 -= 1
                            else:
                                count_dvt1 -= 1
                        img2 = img[1620:2200, 770:832]
                        text = pytesseract.image_to_string(img2, lang = 'vie',config= '--oem 3 --psm 6')
                        lines = text.strip().splitlines()
                        filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
                        sht.range("C9").value = 'Đơn vị tính'
                        for index, line in enumerate(list_dvt, start=10 + count1):  
                            sht[f"C{index}"].value = line 
                        
                        list_sl = []
                        count_sl1 = count
                        img1 = img[897:1620, 1000:1115]
                        img1 = cv2.cvtColor(img1, cv2.COLOR_RGB2GRAY)
                        text = pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6')
                        lines = text.strip().splitlines()
                        filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
                        for i in filtered_lines[:count_sl1]:
                            i = format_string_dg(i)
                            list_sl.append(i)
                            count_sl1 -= 1
                        sht.range("D9").value = 'Số lượng'
                        for index, line in enumerate(list_sl, start=10 + count1):  
                            sht[f"D{index}"].number_format = "@"
                            sht[f"D{index}"].value = line 

                        list_dg = []
                        count_dg1 = count
                        img1 = img[897:1620, 1135:1327]
                        text = pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6')
                        lines = text.strip().splitlines()
                        filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
                        for i in filtered_lines[:count_dg1]:
                            i = format_string_tt(i)
                            list_dg.append(i)
                            count_dg1 -= 1
                        sht.range("E9").value = 'Đơn giá'
                        for index, line in enumerate(list_dg, start=10 + count1):  
                            sht[f"E{index}"].number_format = "@"
                            sht[f"E{index}"].value = line

                        list_tt = []
                        count_tt1 = count
                        img1 = img[897:1620, 1400:1568]
                        text = pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6')
                        lines = text.strip().splitlines()
                        filtered_lines = [line for line in lines if line.strip() != "" and line.strip() != " "]
                        for i in filtered_lines[:count_tt1]:
                            list_tt.append(i)
                            count_tt1 -= 1
                        sht.range("F9").value = 'Thành tiền'
                        for index, line in enumerate(list_tt, start=10+ count1):  
                            sht[f"F{index}"].number_format = "#,##0"
                            text_value = line
                            numeric_value = float(text_value.replace('.', '').replace(',', ''))
                            sht[f"F{index}"].value = numeric_value
                            
                        img1 = img[895:1670, 820:1570]
                        result_2_1 = pytesseract.image_to_data(img1, lang = 'vie+eng', config= 'tessdata',output_type= Output.DICT)
                        filtered_data = [item for item in result_2_1['text'] if item.strip() != '' and item.strip() != ' ']
                        for i, item in enumerate(filtered_data):
                            if item == "Cộng":
                                sht.range("E4").value = 'Cộng tiền hàng'
                                sht.range("F4").number_format = "#,##0"
                                value = filtered_data[i + 3].replace(',', '.')
                                numeric_value = float(value.replace('.', '').replace(',', ''))
                                sht.range("F4").value = numeric_value
                            elif item == "Tiền":
                                sht.range("E5").value = 'Tiền thuế GTGT'
                                sht.range("F5").number_format = "#,##0"
                                value = filtered_data[i + 3].replace(',', '.')
                                if value == "X" or value == "x":
                                    sht.range("F5").value = value
                                else:
                                    numeric_value = float(value.replace('.', '').replace(',', ''))
                                    sht.range("F5").value = numeric_value
                            elif item == "thanh":
                                sht.range("E6").value = 'Tổng tiền thanh toán'
                                sht.range("F6").number_format = "#,##0"
                                value = filtered_data[i + 2].replace(',', '.')
                                numeric_value = float(value.replace('.', '').replace(',', ''))
                                sht.range("F6").value = numeric_value
                else:
                    img_path = os.path.join(path_compare1, myList_compare1[index_1])
                    os.remove(img_path)
                    break

    file_name = f"{so}_KhoaLinh_INV.xlsx"
    wb.save(fr'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Data_Excel\{file_name}')
    excel_file_path = fr'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Data_Excel\{file_name}'  
    columns_bold(excel_file_path, 1,7,1)
    columns_bold(excel_file_path, 1,5,3)
    columns_bold(excel_file_path, 1,4,5)
    rows_bold(excel_file_path, 1, 6, 9)
    wb = xw.Book(excel_file_path)
    change_font_size(wb, 12)
    process_all_worksheets(excel_file_path)

def invoice_pepsico(img,supplier,mst1,address1,consumer,mst2,address2,ms,kh,so,ngaygiao):
    wb = xw.Book()
    sht = wb.sheets.active

    sht.range("A1").value = 'Nhà cung cấp'
    sht.range("B1").value = supplier

    sht.range("A2").value = 'Mã số thuế'
    sht.range("B2").number_format = "@"
    sht.range("B2").value = mst1

    sht.range("A3").value = 'Địa chỉ'
    sht.range("B3").value = address1

    sht.range("A5").value = 'Người mua hàng'
    sht.range("B5").value = consumer

    sht.range("A6").value = 'Mã số thuế'
    sht.range("B6").number_format = "@"
    sht.range("B6").value = mst2

    sht.range("A7").value = 'Địa chỉ'
    sht.range("B7").value = address2

    sht.range("C1").value = 'Mẫu số'
    sht.range("D1").number_format = "@"
    sht.range("D1").value = ms

    sht.range("C2").value = 'Ký hiệu'
    sht.range("D2").number_format = "@"
    sht.range("D2").value = kh

    sht.range("C3").value = 'Số'
    sht.range("D3").number_format = "@"
    sht.range("D3").value = so

    sht.range("C4").value = 'Ngày giao'
    sht.range("D4").number_format = "@"
    sht.range("D4").value = ngaygiao

    #STT
    list_stt = []
    img_stt = img[1098:2200, 98:150]
    data_stt = pytesseract.image_to_string(img_stt, lang = 'eng',config= '--oem 3 --psm 6')
    lines_data_stt = data_stt.strip().splitlines()
    pattern = re.compile(r'^\d+$')
    numbers_stt = [element for element in lines_data_stt if pattern.search(element)]
    for i in numbers_stt:
        list_stt.append(i)
    count = len(list_stt)
    sht.range("A9").value = 'STT'
    for index, line in enumerate(list_stt, start=10):
        sht[f"A{index}"].number_format = "@"  
        sht[f"A{index}"].value = line

    #Tên hàng hóa, dịch vụ
    count_hh1 = count
    img_hh = img[1098:2200, 318:765]
    data_hh = pytesseract.image_to_string(img_hh, lang = 'vie',config= '--oem 3 --psm 6')
    data_hh =  data_hh.replace("NHE","NHF").replace("tỉnh","tinh").replace("muôi","muối")
    lines_data_hh = data_hh.strip().splitlines()
    for index,item in enumerate(lines_data_hh):
        if len(item) <= 20 and index > 0:
            lines_data_hh[index-1] = lines_data_hh[index-1] + " " + lines_data_hh[index]
            lines_data_hh.remove(lines_data_hh[index])
    sht.range("B9").value = "Tên hàng hóa, dịch vụ"
    for index, line in enumerate(lines_data_hh[:count_hh1], start=10):  
        sht[f"B{index}"].value = line 

    #Đơn vị tính, số lượng, đơn giá, thành tiền
    list_dvt = []
    list_sl = []
    list_dg = []
    list_tt = []
    temp_data = []
    img_complex = img[1098:2200, 766:1570]
    data_dvt = pytesseract.image_to_string(img_complex, lang = 'vie',config= '--oem 3 --psm 6')
    lines_data_dvt = data_dvt.strip().splitlines()
    for i in range(0,len(lines_data_dvt[:count])):
        temp_data.append(lines_data_dvt[:count][i].split())
        list_dvt.append(temp_data[i][0])
        list_sl.append(temp_data[i][1])
        list_dg.append(temp_data[i][2])
        list_tt.append(temp_data[i][3])

    sht.range("C9").value = 'Đơn vị tính'
    for index, line in enumerate(list_dvt, start=10):
        sht[f"C{index}"].value = line

    sht.range("D9").value = 'Số lượng'
    for index, line in enumerate(list_sl, start=10):
        sht[f"D{index}"].number_format = "@"
        sht[f"D{index}"].value = line

    sht.range("E9").value = 'Đơn giá'
    for index, line in enumerate(list_dg, start=10):
        sht[f"E{index}"].number_format = "#,##0"
        line = line.replace('.', ',') if '.' in line else line
        line = float(line.replace(',', '')) if ',' in line else float(line)
        sht[f"E{index}"].value = line

    sht.range("F9").value = 'Thành tiền'
    for index, line in enumerate(list_tt, start=10):
        sht[f"F{index}"].number_format = "#,##0"
        line = line.replace('.', ',') if '.' in line else line
        line = float(line.replace(',', '')) if ',' in line else float(line)
        sht[f"F{index}"].value = line
        if line == 0:
            sht[f"E{index}"].value = line

    #Total
    img_total = img[1628:2200, 763:1565]
    data_total = pytesseract.image_to_string(img_total, lang = 'eng',config= '--oem 3 --psm 6')
    lines_data_stt = data_total.strip().splitlines()
    for item in lines_data_stt:
        if "(Total amount)" in item:
            colon_index = item.index(':')
            untax = item[colon_index + 1:].strip()
        if "(VAT amount)" in item:
            colon_index = item.index(':')
            tax = item[colon_index + 1:].strip()
        if "(Total amount due)" in item:
            colon_index = item.index(':')
            total = item[colon_index + 1:].strip()
    
    value = untax.replace('.', ',') if '.' in untax else untax
    sht.range("E1").value = 'Cộng tiền hàng chưa có thuế GTGT'
    sht.range("F1").number_format = "#,##0"
    text_value = value
    numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
    sht.range("F1").value = numeric_value

    value = tax.replace('.', ',') if '.' in tax else tax
    sht.range("E2").value = 'Tiền thuế GTGT'
    sht.range("F2").number_format = "#,##0"
    if value.replace(',', '').replace('.', '').isdigit():
        text_value = value
        numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
    else:
        # Nếu value không phải là số, giữ nguyên value
        text_value = value
        numeric_value = text_value
    sht.range("F2").value = numeric_value

    value = total.replace('.', ',') if '.' in total else total
    sht.range("E3").value = 'Tổng cộng tiền thanh toán'
    sht.range("F3").number_format = "#,##0"
    text_value = value
    numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
    sht.range("F3").value = numeric_value

    path_compare1 = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Compare_Invoice'
    path_source1 = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Images_Invoice'
    orb = cv2.ORB_create(nfeatures = 1000)
    images_compare1 = []
    images_source1 = []
    myList_compare1 = os.listdir(path_compare1)
    mylist_source1 = os.listdir(path_source1)
    for img1 in myList_compare1:
        images_compare1 = cv2.imread(f'{path_compare1}/{img1}')
        images_compare1 = cv2.cvtColor(images_compare1,cv2.COLOR_BGR2GRAY)
        kp1, des1 = orb.detectAndCompute(images_compare1,None)
        matchList1 = []
        count1 = count
        for img2 in mylist_source1:
            index_1 = myList_compare1.index(img1)
            images_source1 = cv2.imread(f'{path_source1}/{img2}')
            kp2, des2 = orb.detectAndCompute(images_source1,None)
            bf = cv2.BFMatcher()
            matches = bf.knnMatch(des1,des2,k=2)
            good = []
            for m,n in matches:
                if m.distance < 0.75*n.distance:
                    good.append([m])
            matchList1.append(len(good))
        if max(matchList1) > 450:
            img_path = os.path.join(path_compare1, myList_compare1[index_1])
            list_stt1 = []
            img_stt = img[1098:2200, 98:150]
            data_stt = pytesseract.image_to_string(img_stt, lang = 'eng',config= '--oem 3 --psm 6')
            lines_data_stt = data_stt.strip().splitlines()
            pattern = re.compile(r'^\d+$')
            numbers_stt = [element for element in lines_data_stt if pattern.search(element)]
            for i in numbers_stt:
                list_stt1.append(i)
            count = len(list_stt1)
            if count == 0:
                img_total = img[1628:2200, 763:1565]
                data_total = pytesseract.image_to_string(img_total, lang = 'eng',config= '--oem 3 --psm 6')
                lines_data_stt = data_total.strip().splitlines()
                for item in lines_data_stt:
                    if "(Total amount)" in item:
                        colon_index = item.index(':')
                        untax = item[colon_index + 1:].strip()
                    if "(VAT amount)" in item:
                        colon_index = item.index(':')
                        tax = item[colon_index + 1:].strip()
                    if "(Total amount due)" in item:
                        colon_index = item.index(':')
                        total = item[colon_index + 1:].strip()
                
                value = untax.replace('.', ',') if '.' in untax else untax
                sht.range("E1").value = 'Cộng tiền hàng chưa có thuế GTGT'
                sht.range("F1").number_format = "#,##0"
                text_value = value
                numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
                sht.range("F1").value = numeric_value

                value = tax.replace('.', ',') if '.' in tax else tax
                sht.range("E2").value = 'Tiền thuế GTGT'
                sht.range("F2").number_format = "#,##0"
                if value.replace(',', '').replace('.', '').isdigit():
                    text_value = value
                    numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
                else:
                    # Nếu value không phải là số, giữ nguyên value
                    text_value = value
                    numeric_value = text_value
                sht.range("F2").value = numeric_value

                value = total.replace('.', ',') if '.' in total else total
                sht.range("E3").value = 'Tổng cộng tiền thanh toán'
                sht.range("F3").number_format = "#,##0"
                text_value = value
                numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
                sht.range("F3").value = numeric_value
            else:
                sht.range("A9").value = 'STT'
                for index, line in enumerate(list_stt1, start=10 + count1):
                    sht[f"A{index}"].number_format = "@"  
                    sht[f"A{index}"].value = line

                #Tên hàng hóa, dịch vụ
                img_hh = img[1098:2200, 318:765]
                #cv2.imshow("img_hh",img_hh)
                data_hh = pytesseract.image_to_string(img_hh, lang = 'vie',config= '--oem 3 --psm 6')
                data_hh =  data_hh.replace("NHE","NHF").replace("tỉnh","tinh").replace("muôi","muối")
                lines_data_hh = data_hh.strip().splitlines()
                for index,item in enumerate(lines_data_hh):
                    if len(item) <= 20 and index > 0:
                        lines_data_hh[index-1] = lines_data_hh[index-1] + " " + lines_data_hh[index]
                        lines_data_hh.remove(lines_data_hh[index])
                sht.range("B9").value = "Tên hàng hóa, dịch vụ"
                for index, line in enumerate(lines_data_hh[:count], start=10 + count1):  
                    sht[f"B{index}"].value = line 

                #Đơn vị tính, số lượng, đơn giá, thành tiền
                list_dvt1 = []
                list_sl1 = []
                list_dg1 = []
                list_tt1 = []
                temp_data1 = []
                img_complex = img[1098:2200, 766:1570]
                data_dvt = pytesseract.image_to_string(img_complex, lang = 'vie',config= '--oem 3 --psm 6')
                lines_data = data_dvt.strip().splitlines()
                for i in range(0,len(lines_data[:count])):
                    temp_data1.append(lines_data[:count][i].split())
                    list_dvt1.append(temp_data1[i][0])
                    list_sl1.append(temp_data1[i][1])
                    list_dg1.append(temp_data1[i][2])
                    list_tt1.append(temp_data1[i][3])

                sht.range("C9").value = 'Đơn vị tính'
                for index, line in enumerate(list_dvt1, start=10+count1):
                    sht[f"C{index}"].value = line

                sht.range("D9").value = 'Số lượng'
                for index, line in enumerate(list_sl1, start=10+count1):
                    sht[f"D{index}"].number_format = "@"
                    sht[f"D{index}"].value = line

                sht.range("E9").value = 'Đơn giá'
                for index, line in enumerate(list_dg1, start=10+count1):
                    sht[f"E{index}"].number_format = "#,##0"
                    line = line.replace('.', ',') if '.' in line else line
                    line = float(line.replace(',', '')) if ',' in line else float(line)
                    sht[f"E{index}"].value = line

                sht.range("F9").value = 'Thành tiền'
                for index, line in enumerate(list_tt1, start=10+count1):
                    sht[f"F{index}"].number_format = "#,##0"
                    line = line.replace('.', ',') if '.' in line else line
                    line = float(line.replace(',', '')) if ',' in line else float(line)
                    sht[f"F{index}"].value = line
                    if line == 0:
                        sht[f"E{index}"].value = line

                #Total
                img_total = img[1628:2200, 763:1565]
                data_total = pytesseract.image_to_string(img_total, lang = 'eng',config= '--oem 3 --psm 6')
                lines_data_stt = data_total.strip().splitlines()
                for item in lines_data_stt:
                    if "(Total amount)" in item:
                        colon_index = item.index(':')
                        untax = item[colon_index + 1:].strip()
                    if "(VAT amount)" in item:
                        colon_index = item.index(':')
                        tax = item[colon_index + 1:].strip()
                    if "(Total amount due)" in item:
                        colon_index = item.index(':')
                        total = item[colon_index + 1:].strip()
                
                value = untax.replace('.', ',') if '.' in untax else untax
                sht.range("E1").value = 'Cộng tiền hàng chưa có thuế GTGT'
                sht.range("F1").number_format = "#,##0"
                text_value = value
                numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
                sht.range("F1").value = numeric_value

                value = tax.replace('.', ',') if '.' in tax else tax
                sht.range("E2").value = 'Tiền thuế GTGT'
                sht.range("F2").number_format = "#,##0"
                if value.replace(',', '').replace('.', '').isdigit():
                    text_value = value
                    numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
                else:
                    # Nếu value không phải là số, giữ nguyên value
                    text_value = value
                    numeric_value = text_value
                sht.range("F2").value = numeric_value

                value = total.replace('.', ',') if '.' in total else total
                sht.range("E3").value = 'Tổng cộng tiền thanh toán'
                sht.range("F3").number_format = "#,##0"
                text_value = value
                numeric_value = float(text_value.replace(',', '')) if ',' in text_value else float(text_value)
                sht.range("F3").value = numeric_value
        else:
            img_path = os.path.join(path_compare1, myList_compare1[index_1])
            os.remove(img_path)
            break
    file_name = f"{so}_Pepsico_INV.xlsx"
    wb.save(fr'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Data_Excel\{file_name}')
    excel_file_path = fr'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Data_Excel\{file_name}'  
    columns_bold(excel_file_path, 1,7,1)
    columns_bold(excel_file_path, 1,5,3)
    columns_bold(excel_file_path, 1,4,5)
    rows_bold(excel_file_path, 1, 6, 9)
    wb = xw.Book(excel_file_path)
    change_font_size(wb, 12)
    process_all_worksheets(excel_file_path)

def main():
    start = time()
    pdf_to_png_INV()
    verify_executed_Cocacola = False
    verify_executed_LinhKhoa = False
    verify_executed_Pepsico = False
    path_compare_INV = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Compare_Invoice'
    path_source_INV = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Images_Invoice'
    orb = cv2.ORB_create(nfeatures = 1000)
    images_compare_INV = []
    images_source_INV = []
    myList_compare_INV = os.listdir(path_compare_INV)
    print(myList_compare_INV)
    mylist_source_INV = os.listdir(path_source_INV)
    print('Số lượng hóa đơn được trích xuất:',len(myList_compare_INV))
    for index in range(len(myList_compare_INV)):
        if len(myList_compare_INV) > 0:
            img1 = myList_compare_INV[0]
            print(img1)
            images_compare_INV = cv2.imread(f'{path_compare_INV}\{img1}')
            images_compare_INV = cv2.cvtColor(images_compare_INV,cv2.COLOR_BGR2GRAY)
            kp1, des1 = orb.detectAndCompute(images_compare_INV,None)
            matchList = []
            for img2 in mylist_source_INV:
                index_1 = myList_compare_INV.index(img1)
                print(index_1)
                images_source_INV = cv2.imread(f'{path_source_INV}/{img2}')
                kp2, des2 = orb.detectAndCompute(images_source_INV,None)
                finalVal = -1
                bf = cv2.BFMatcher()
                matches = bf.knnMatch(des1,des2,k=2)
                good = []
                for m,n in matches:
                    if m.distance < 0.75*n.distance:
                        good.append([m])  
                matchList.append(len(good))
                finalVal = -1 
                print(matchList)
                for i, num_matches in enumerate(matchList):
                    if num_matches > 450:
                        finalVal = i
                        break
            if finalVal == 0:
                img_path = os.path.join(path_compare_INV, myList_compare_INV[index_1])
                img = cv2.imread(img_path)
                img1 = img[116:303,388:1235]
                string1 =  pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6 ')
                lines1 = string1.strip().splitlines()
                data_cleaned1 = [item for item in lines1 if item != '']

                supplier = data_cleaned1[0] + ' '+ data_cleaned1[1]

                colon_index = data_cleaned1[2].index(':')
                mst1 = data_cleaned1[2][colon_index + 1:].strip()

                colon_index = data_cleaned1[3].index(':')
                address1 = data_cleaned1[3][colon_index + 1:].strip()

                address1 = address1 + " " + data_cleaned1[4]
                address1 = address1.replace("Sô","Số")

                img2 = img[450:671,93:1580]
                string2 =  pytesseract.image_to_string(img2, lang = 'vie',config= '--oem 3 --psm 6 ')
                lines2 = string2.strip().splitlines()
                data2 = []
                for line in lines2:
                    if ':' in line:
                        colon_index = line.index(':')
                        data = line[colon_index + 1:].strip()
                        data2.append(data)
                consumer = data2[1]

                address2 = data2[2]

                mst2 = data2[3]

                img3 = img[116:250,1235:1520]
                string3 =  pytesseract.image_to_string(img3, lang = 'eng',config= '--oem 3 --psm 6 ')
                lines3 = string3.strip().splitlines()
                data3 = []
                for i in range(0,len(lines3)):
                    colon_index = lines3[i].index(':')
                    data = lines3[i][colon_index + 1:].strip()
                    data3.append(data)

                ms = data2[0]

                kh = data3[0]

                so = data3[1]

                ngaygiao = data3[2]
                # while True:
                #         if not verify_executed_Pepsico:  
                #             text_verify = verify_invoice(mst1,ms,kh,so)
                #             verify_executed_Cocacola = True  
                #         else:
                #             break
                #         if text_verify == "NNT tạm nghỉ kinh doanh có thời hạn":
                #             print("Công ty Pepsico đã thông báo về việc tạm ngừng hoạt động có thời hạn và được cơ quan có thẩm quyền chấp thuận")
                #             print(myList_compare_INV[index_1])
                #             break
                #         elif text_verify == "NNT ngừng hoạt động nhưng chưa hoàn thành thủ tục đóng MST":
                #             print("Doanh nghiệp Pepsico không hoàn thành các nghĩa vụ thuế")
                #             print(myList_compare_INV[index_1])
                #             break
                #         elif text_verify == "NNT không hoạt động tại địa chỉ đã đăng ký":
                #             print("Công ty Pepsico đang tra cứu đã bị cơ quan thuế quản lý khóa mã số thuế do doanh nghiệp không hoạt động tại địa điểm như đã đăng ký trên Giấy chứng nhận đăng ký kinh doanh")
                #             print(myList_compare_INV[index_1])
                #             break
                #         elif text_verify == "NNT đang hoạt động (đã được cấp GCN ĐKT)":
                #             print("Nhà cung cấp Pepsico hợp lệ")
                #             break
                os.remove(img_path)
                invoice_pepsico(img,supplier,mst1,address1,consumer,mst2,address2,ms,kh,so,ngaygiao)
                path_compare_INV = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Compare_Invoice'
                myList_compare_INV = os.listdir(path_compare_INV)
            elif finalVal == 3: 
                img_path = os.path.join(path_compare_INV, myList_compare_INV[index_1])
                img = cv2.imread(img_path)
                img1 = img[40:250, 95:1560]
                string1 =  pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6 ')
                lines1 = string1.strip().splitlines()
                data_cleaned1 = [item for item in lines1 if item != '']
                supplier = data_cleaned1[0]

                colon_index = lines1[-1].index(':')
                mst1 = lines1[-1][colon_index + 1:].strip() 
                mst1_1 = mst1.split()

                colon_index = mst1.index(':')
                ms = mst1[colon_index + 1:].strip() 
                mst1 = mst1_1[0]

                address1 = data_cleaned1[2]

                img2 = img[280:450,1107:1550]
                string2 =  pytesseract.image_to_string(img2, lang = 'vie',config= '--oem 3 --psm 6 ')
                lines2 = string2.strip().splitlines()
                temp_lines2 = lines2[1]
                temp_lines2 = temp_lines2.split()

                kh = temp_lines2[0]

                so = temp_lines2[1]

                colon_index = lines2[-1].index(':')
                ngaygiao = lines2[-1][colon_index + 1:].strip() 

                img3 = img[480:620,494:1218]
                string3 =  pytesseract.image_to_string(img3, lang = 'vie',config= '--oem 3 --psm 6 preserve_interword_spaces=1')
                lines3 = string3.strip().splitlines()
                consumer = lines3[1]

                img4 = img[682:775,1216:1555]
                string4 =  pytesseract.image_to_string(img4, lang = 'eng',config= '--oem 3 --psm 6')
                lines4 = string4.strip().splitlines()
                mst2 = lines4[0]

                img5 = img[780:980,105:1210]
                dcch = pytesseract.image_to_string(img5, lang = 'vie',config= '--oem 3 --psm 6')
                dcch = dcch.strip().splitlines()
                dcch = [line for line in dcch if line.strip()]
                if len(dcch) > 4:
                    dcch[3] = dcch[3] + " " + dcch[4]
                    dcch.remove(dcch[4])
                for line in dcch:
                    if "Việt Nam" in line:
                        address2 = line.replace("TP", "Thành phố").replace(" ,",",").replace("Kios","Kiost").replace("Phố","phố").replace("Hò","Hồ")

                img6 = img[1910:2030,220:710]
                string6 =  pytesseract.image_to_string(img6, lang = 'eng',config= '--oem 3 --psm 6 ')
                lines6 = string6.strip().splitlines()
                colon_index = lines6[-1].index(':')
                # while True:
                #         if not verify_executed_Cocacola:  
                #             text_verify = verify_invoice(mst1,ms,kh,so)
                #             verify_executed_Cocacola = True  
                #         else:
                #             break
                #         if text_verify == "NNT tạm nghỉ kinh doanh có thời hạn":
                #             print("Công ty Cocacola đã thông báo về việc tạm ngừng hoạt động có thời hạn và được cơ quan có thẩm quyền chấp thuận")
                #             print(myList_compare_INV[index_1])
                #             break
                #         elif text_verify == "NNT ngừng hoạt động nhưng chưa hoàn thành thủ tục đóng MST":
                #             print("Doanh nghiệp Cocacola không hoàn thành các nghĩa vụ thuế")
                #             print(myList_compare_INV[index_1])
                #             break
                #         elif text_verify == "NNT không hoạt động tại địa chỉ đã đăng ký":
                #             print("Công ty Cocacola đang tra cứu đã bị cơ quan thuế quản lý khóa mã số thuế do doanh nghiệp không hoạt động tại địa điểm như đã đăng ký trên Giấy chứng nhận đăng ký kinh doanh")
                #             print(myList_compare_INV[index_1])
                #             break
                #         elif text_verify == "NNT đang hoạt động (đã được cấp GCN ĐKT)":
                #             print("Nhà cung cấp Cocacola hợp lệ")
                #             break
                os.remove(img_path)
                invoice_cocacola(img,supplier,mst1,address1,consumer,mst2,address2,ms,kh,so,ngaygiao)
                path_compare_INV = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Compare_Invoice'
                myList_compare_INV = os.listdir(path_compare_INV)
            elif finalVal == 5: 
                img_path = os.path.join(path_compare_INV, myList_compare_INV[index_1])
                img = cv2.imread(img_path)
                img[579:745,1379:1540] = 255
                img1 = img[270:791, 66:1534]
                string1 =  pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6 preserve_interword_spaces=1')
                lines1 = string1.strip().splitlines()
                data_cleaned1 = [item for item in lines1 if item != '']
                supplier = data_cleaned1[0]

                colon_index = data_cleaned1[1].index(':')
                mst1 = data_cleaned1[1][colon_index + 1:].strip()

                colon_index = data_cleaned1[2].index(':')
                address1 = data_cleaned1[2][colon_index + 1:].strip()

                colon_index = data_cleaned1[7].index(':')
                consumer = data_cleaned1[7][colon_index + 1:].strip()

                colon_index = data_cleaned1[8].index(':')
                mst2 = data_cleaned1[8][colon_index + 1:].strip()

                colon_index = data_cleaned1[9].index(':')
                address2 = data_cleaned1[9][colon_index + 1:].strip()
                address2 = address2.replace("Mguyễn","Nguyễn")

                img2 = img[70:200,300:1600]
                string2 =  pytesseract.image_to_string(img2, lang = 'eng',config= '--oem 3 --psm 6 preserve_interword_spaces=1')
                lines2 = string2.strip().splitlines()
                colon_index = lines2[0].index(':')
                kh = lines2[0][colon_index + 1:].strip()

                temp_lines2 = lines2[1].split()
                ngaygiao = temp_lines2[1] + "/" + temp_lines2[3] + "/" + temp_lines2[5]

                colon_index = lines2[1].index(':')
                so = lines2[1][colon_index + 1:].strip()

                img3 = img[190:247,614:1135]
                img3 = cv2.cvtColor(img3,cv2.COLOR_BGR2GRAY)
                img3 = cv2.resize(img3,None,fx=1.24,fy=1.24,interpolation=cv2.INTER_BITS)
                gamma = 2
                img3 = np.uint8(np.power(img3 / float(np.max(img3)), gamma) * 255)
                ms = pytesseract.image_to_string(img3,lang= "vie", config='--oem 3 --psm 7')
                ms = ms.replace("&","").replace("O","0").replace("I","1").replace("Q","0").replace("68B","6B").replace("383","3B3").replace("984D","98AD").replace("340","3A0").replace("085A","08BA").replace("946","9A6").replace("984","98A").replace("346","3A6")
                ms = re.sub(r'[^a-fA-F0-9]', '', ms)
                if not ms.startswith("00"):
                    ms = "0" + ms
                # while True:
                #         if not verify_executed_LinhKhoa:  
                #             text_verify = verify_invoice(mst1,ms,kh,so)
                #             verify_executed_LinhKhoa = True  
                #         else:
                #             break
                #         if text_verify == "NNT tạm nghỉ kinh doanh có thời hạn":
                #             print("Công ty Linh Khoa đã thông báo về việc tạm ngừng hoạt động có thời hạn và được cơ quan có thẩm quyền chấp thuận")
                #             print(myList_compare_INV[index_1])
                #             break
                #         elif text_verify == "NNT ngừng hoạt động nhưng chưa hoàn thành thủ tục đóng MST":
                #             print("Doanh nghiệp Linh Khoa không hoàn thành các nghĩa vụ thuế")
                #             print(myList_compare_INV[index_1])
                #             break
                #         elif text_verify == "NNT không hoạt động tại địa chỉ đã đăng ký":
                #             print("Công ty Linh Khoa đang tra cứu đã bị cơ quan thuế quản lý khóa mã số thuế do doanh nghiệp không hoạt động tại địa điểm như đã đăng ký trên Giấy chứng nhận đăng ký kinh doanh")
                #             print(myList_compare_INV[index_1])
                #             break
                #         elif text_verify == "NNT đang hoạt động (đã được cấp GCN ĐKT)":
                #             print("Nhà cung cấp Linh Khoa hợp lệ")
                #             break
                os.remove(img_path)
                invoice_linhkhoa(img,supplier,mst1,address1,consumer,mst2,address2,ms,kh,so,ngaygiao)
                path_compare_INV = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Compare_Invoice'
                myList_compare_INV = os.listdir(path_compare_INV)
            else:
                path_compare_INV = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Compare_Invoice'
                img_path = os.path.join(path_compare_INV, myList_compare_INV[index_1])
                convert_image_to_pdf(img_path)
                os.remove(img_path)
                myList_compare_INV = os.listdir(path_compare_INV)
    folder_path = r"SourcecodeOCR\Source_Compare_Invoice"
    delete_img_files(folder_path)
    file_count = len(os.listdir("SourcecodeOCR\File_PDF_Invoice"))
    print(f"Số lượng hóa đơn không thể trích xuất là: {file_count}")
    for filename in os.listdir("SourcecodeOCR\File_PDF_Invoice"):
        print(filename)
    Excel_Jusified_INV()
    folder_pdf = r"SourcecodeOCR\File_PDF_Invoice"
    delete_pdf_files(folder_pdf)

    print("Quá trình trích xuất hóa đơn hoàn tất")
    print(time()-start)

if __name__ == "__main__":
    main() 

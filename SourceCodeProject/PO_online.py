import pytesseract 
from pytesseract import Output
import cv2
import os
import xlwings as xw
from pdf2image import convert_from_path
import re
from datetime import datetime
from time import sleep,time
import img2pdf
import glob
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Font,numbers,NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl    
from pathlib import Path
from random import *
import pandas as pd
import comtypes.client

pytesseract.pytesseract.tesseract_cmd =r'H:\\APP UNIVERSITY\\CODE PYTHON\\Tesseract-ocr\\tesseract.exe'
def clear_cells(sheet, start_col, end_col, row_number):
        for col in range(start_col, end_col + 1):
            sheet.cell(row=row_number, column=col).value = ''

def format_number_with_commas(number_str):
        if number_str is None:
            return ""
        
        if isinstance(number_str, int):
            number_str = str(number_str)
        
        number_str = re.sub(r'[,]', '', number_str)  # Loại bỏ dấu phẩy
        try:
            number = int(number_str)
            formatted_number = f"{number:,}"
            return formatted_number
        except ValueError:
            return number_str
        
def convert_string_to_number(s):
        try:
            s = s.replace(',', '')  # Loại bỏ dấu phẩy (,) trong chuỗi
            return float(s)  # Chuyển đổi thành số thực
        except ValueError:
            return None  # Không thể chuyển đổi thành số
        
def Excel_Jusified_PO():
    location = 'H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_PO\\*.xlsx'
    excel_files = glob.glob(location)
    output_path = 'H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_PO\\Total_PO.xlsx'

    workbook = Workbook()
    for excel_file in excel_files:
        sheet_name = os.path.basename(excel_file)[:31]
        df1 = pd.read_excel(excel_file, engine='openpyxl')
        df1.fillna(value='', inplace=True)
        sheet = workbook.create_sheet(sheet_name)
        for row in dataframe_to_rows(df1, index=False, header=True):
            sheet.append(row)
    # Xóa trang mặc định
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)
    # Lưu Workbook vào tệp mới
    workbook.save(output_path)
    output_excel_file = 'H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_PO\\Total_PO.xlsx'
    excel = openpyxl.load_workbook(output_excel_file)
    sheet_names = excel.sheetnames
    font = Font(name='Times New Roman', size=12)
    sheet_counter = 0 
    for sheet_name in sheet_names:
        sheet = excel[sheet_name]
        max_widths = []
        clear_cells(sheet, 3, 4, 1) 
        for row in sheet.iter_rows():
            for i, cell in enumerate(row):
                cell.font = font
                cell_value = str(cell.value)
                if i >= len(max_widths):
                    max_widths.append(len(cell_value))
                else:
                    max_widths[i] = max(max_widths[i], len(cell_value))
        for i, column_width in enumerate(max_widths, start=1):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = column_width + 4 #2  
        for row in sheet.iter_rows(min_row=1, min_col=1, max_col=len(max_widths)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')
        for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')
        for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')
        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')
        for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')        
        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)
                
        for row in sheet.iter_rows(min_row=8, max_row=8, min_col=1, max_col=7):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)

        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=5, max_col=5):
            for cell in row:
                cell.font = Font(bold=True,name='Times New Roman', size=12)
        for row in sheet.iter_rows(min_row=4, max_row=6, min_col=6, max_col=6):
            for cell in row:
                cell_value = format_number_with_commas(cell.value)
                cell.value = cell_value
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, min_col=5, max_col=6):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell_value = format_number_with_commas(str(cell.value))
                    cell.value = cell_value
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        sum = 0
        for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, min_col=6, max_col=6):
            for cell in row:
                cell_value = str(cell.value)
                number_value = convert_string_to_number(cell_value)
                
                if number_value is not None:
                    sum += number_value
        cell_F4 = sheet['F4']
        value_F4 = convert_string_to_number(str(cell_F4.value))
        if sum == value_F4:
            None
        else:
            print(f"Sheet {sheet_counter + 1} có tổng thành tiền khác với cộng tiền hàng chưa thuế.")
        sheet_counter += 1
    # Lưu lại tệp Excel đã cập nhật
    output_filename_with_time = f'Total_PO_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    output_excel_file_with_time = os.path.join("H:\\APP UNIVERSITY\\CODE PYTHON\\OpenCV-Master-Computer-Vision-in-Python\\SourcecodeOCR\\Data_PO", output_filename_with_time)
    excel.save(output_excel_file_with_time)
    total_PO_path = os.path.join("H:/APP UNIVERSITY/CODE PYTHON/OpenCV-Master-Computer-Vision-in-Python/SourcecodeOCR/Data_PO", "Total_PO.xlsx")
    os.remove(total_PO_path)

    excel = openpyxl.load_workbook(output_excel_file_with_time)
    new_sheet_name = "Data_Total"
    new_sheet = excel.create_sheet(new_sheet_name)
    new_sheet['A12'] = "STT"
    new_sheet['B12'] = "Ký hiệu PO"
    new_sheet['C12'] = "Ngày giao hàng"
    new_sheet['D12'] = "Tên người bán"
    new_sheet['E12'] = "Mặt hàng"
    new_sheet['F12'] = "Doanh số mua chưa có thuế"
    new_sheet['G12'] = "Thuế GTGT đủ điều kiện khấu trừ thuế"
    new_sheet['H12'] = "Ghi chú"
    for column in new_sheet.iter_cols(min_col=1, max_col=8, min_row=12, max_row=12):
        for cell in column:
            cell.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')
            cell.font = Font(name='Times New Roman', size=12,bold=True)
    column_widths = {'A': 5,'B': 15,'C': 10,'D': 85,'E': 40,'F': 15,'G': 15,'H': 50}
    for column, width in column_widths.items():
        new_sheet.column_dimensions[column].width = width

    excel.save(output_excel_file_with_time)
    count_sheet = 0
    sheets_with_max_row_9 = []
    for sheet_name in excel.sheetnames[:-1]:  
        sheet = excel[sheet_name]
        new_row = len(new_sheet['A'])+1  # Bắt đầu từ hàng 13
        data_to_copy = sheet['F1'].value
        new_sheet.cell(row=new_row, column=2, value=data_to_copy)
        data_to_copy = sheet['B6'].value
        new_sheet.cell(row=new_row, column=3, value=data_to_copy)
        data_to_copy = sheet['B1'].value
        new_sheet.cell(row=new_row, column=4, value=data_to_copy)
        data_to_copy = sheet['F4'].value
        new_sheet.cell(row=new_row, column=6, value=data_to_copy)
        data_to_copy = sheet['F5'].value
        new_sheet.cell(row=new_row, column=7, value=data_to_copy)
        count_sheet +=1
        if sheet.max_row == 9:
            sheets_with_max_row_9.append(sheet_name)
            count_note = count_sheet
            if sheets_with_max_row_9:
                target_sheet_name = sheets_with_max_row_9[0] 
                target_worksheet = excel[target_sheet_name]
                value_at_B9 = target_worksheet['B19'].value
                last_worksheet_name = excel.sheetnames[-1]
                last_worksheet = excel[last_worksheet_name]
                last_worksheet[f'G{12 + count_note}'] = value_at_B9
                excel.save(output_excel_file_with_time)
                sheets_with_max_row_9.pop(0)

    for row in range(13, new_sheet.max_row + 1):
        cell_value = new_sheet.cell(row=row, column=5).value
        if ((new_sheet.cell(row=row, column=4).value == "CÔNG TY TRÁCH NHIỆM HỮU HẠN NƯỚC GIẢI KHÁT COCA-COLA VIỆT NAM") and (cell_value == None)) or ((new_sheet.cell(row=row, column=4).value == "CHI NHÁNH CÔNG TY TNHH NƯỚC GIẢI KHÁT COCA-COLA VIỆT NAM TẠI THÀNH PHỐ ĐÀ NẴNG") and (cell_value == None)):
            new_sheet.cell(row=row, column=5, value="Chi phí mua đồ uống giải khát")
        # elif new_sheet.cell(row=row, column=4).value == "CÔNG TY TNHH THỰC PHẨM LINH KHOA" and cell_value == None:
        #     new_sheet.cell(row=row, column=5, value="Chi phí mua thực phẩm")

    for column in new_sheet.iter_cols(min_col=1, max_col=5, min_row=13, max_row=new_sheet.max_row):
        for cell in column:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for index, number in enumerate(range(1, count_sheet + 1)):
        destination_row = 13 + index
        new_sheet.cell(row=destination_row, column=1, value=number)
   
    number_style = NamedStyle(name='number_style', number_format='#,##0')
    for row in new_sheet.iter_rows(min_row=13, max_row=new_sheet.max_row, min_col=6, max_col=7):
        for cell in row:
            cell.style = number_style
    for column in new_sheet.iter_cols(min_col=6, max_col=7, min_row=13, max_row=new_sheet.max_row):
        for cell in column:
            cell.alignment = Alignment(horizontal='right', vertical='center')
    for column in new_sheet.iter_cols(min_col=1, max_col=8, min_row=13, max_row=new_sheet.max_row):
        for cell in column:
            cell.font = Font(name='Times New Roman', size=11)

    output_folder = "H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Data_PO_Final"
    output_excel_file_with_time1 = os.path.join(output_folder, output_filename_with_time)
    excel.save(output_excel_file_with_time1)
    workbook = load_workbook(output_excel_file_with_time1)
    worksheet_names = workbook.sheetnames
    new_names = [f"PO{i+1}" for i in range(len(worksheet_names)-1)]
    for i in range(len(worksheet_names)-1):
        worksheet = workbook[worksheet_names[i]]
        worksheet.title = new_names[i]
    workbook.save(output_excel_file_with_time1)
    os.remove(output_excel_file_with_time)

def pdf_to_png_PO():
    poppler_path = r'H:/OCR/Popler/poppler-23.07.0/Library/bin'
    pdf_folder = r"H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR/File_PDF_PO"
    saving_folder = r"H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Compare_PO"
    os.makedirs(saving_folder, exist_ok=True)
    for pdf_filename in os.listdir(pdf_folder):
        if pdf_filename.lower().endswith(".pdf"):
            pdf_path = os.path.join(pdf_folder, pdf_filename)
            pages = convert_from_path(pdf_path=pdf_path, poppler_path=poppler_path)
            for c, page in enumerate(pages, start=1):
                img_name = f"{os.path.splitext(pdf_filename)[0]}.png"
                img_path = os.path.join(saving_folder, img_name)
                page.save(img_path, "png")
        os.remove(pdf_path)

def make_rows_bold(excel_file_path, start_column, end_column, row):
    wb = xw.Book(excel_file_path)
    sheet = wb.sheets.active
    for col in range(start_column, end_column + 1):
        sheet.range(row, col).api.Font.Bold = True
    wb.save()

def make_columns_bold(excel_file_path, start_row,end_row,col):
    wb = xw.Book(excel_file_path)
    sheet = wb.sheets.active
    for row in range(start_row, end_row + 1):
        sheet.range(row, col).api.Font.Bold = True
    wb.save()

def change_font_size_and_save(wb, font_size):
    for sheet in wb.sheets:
        for row in sheet.used_range.rows:
            for cell in row:
                cell.api.Font.Size = font_size
    wb.save()

def fit_column_widths_for_one_sheet(sheet):
    cell_widths = {}
    used_range = sheet.used_range
    for col in range(1, used_range.columns.count + 1):
        for row in range(1, used_range.rows.count + 1):
            cell_value = used_range(row, col).value
            if cell_value:
                if str(cell_value)[0] == "=":
                    continue
                cell_widths[col] = max(
                    (cell_widths.get(col, 0), len(str(cell_value))+4)
                )
    for col, column_width in cell_widths.items():
        column_width = str(column_width)
        sheet.range((1, col), (used_range.rows.count, col)).column_width = column_width

def process_all_worksheets(excel_file_path):
    wb = xw.Book(excel_file_path)
    for sheet in wb.sheets:
        fit_column_widths_for_one_sheet(sheet)
        sheet.range("A8:F8").api.HorizontalAlignment = -4108  # -4108 tương ứng với giá trị xlCenter trong Excel
        sheet.range("A8:A13").api.HorizontalAlignment = -4108
        sheet.range("A1:B6").api.HorizontalAlignment = -4131
        sheet.range("C8:E13").api.HorizontalAlignment = -4108
        sheet.range("F1:F6").api.HorizontalAlignment = -4152
        sheet.range("F9:F14").api.HorizontalAlignment = -4152
    wb.save()
    wb.close()

#Convert word to pdf
def convert_word_to_pdf_single(word_path, pdf_path):
    # Tạo đối tượng Word
    word = comtypes.client.CreateObject("Word.Application")
    word.Visible = False

    # Mở tệp Word
    in_file = word.Documents.Open(word_path)

    # Lưu tệp Word thành tệp PDF
    in_file.SaveAs(pdf_path, FileFormat=17)  # FileFormat=17 tương ứng với định dạng PDF
    in_file.Close()
    # Đóng ứng dụng Word
    word.Quit()

def PO_cocacola(img,vendor,address1,consumer,ten_ch,address2,ngay_giaohang,no,notes):
    wb = xw.Book()
    sht = wb.sheets.active

    sht.range("A1").value = 'Nhà cung cấp'
    sht.range("B1").value = vendor

    sht.range("A2").value = 'Địa chỉ'
    sht.range("B2").number_format = "@"
    sht.range("B2").value = address1

    sht.range("A3").value = 'Người mua hàng'
    sht.range("B3").value = consumer

    sht.range("A4").value = 'Tên cửa hàng'
    sht.range("B4").number_format = "@"
    sht.range("B4").value = ten_ch

    sht.range("A5").value = 'Địa chỉ cửa hàng'
    sht.range("B5").number_format = "@"
    sht.range("B5").value = address2

    sht.range("A6").value = 'Ngày giao hàng '
    sht.range("B6").number_format = "@"
    sht.range("B6").value = ngay_giaohang

    sht.range("E1").value = 'Số hóa đơn'
    sht.range("F1").number_format = "@"
    sht.range("F1").value = no

    sht.range("E2").value = 'Ghi chú'
    sht.range("F2").number_format = "@"
    sht.range("F2").value = notes

    sht.range("E3").value = 'Check'
    sht.range("F3").number_format = "@"

    list_stt = []
    img1 = img[720:1654, 60:115]
    text1 = pytesseract.image_to_string(img1, lang = 'eng',config= '--oem 3 --psm 6')
    lines1 = text1.strip().splitlines()
    pattern = re.compile(r'^\d+$')
    elements_with_numbers = [element for element in lines1 if pattern.search(element)]
    for i in elements_with_numbers:
        list_stt.append(i)
    count = len(list_stt)
    sht.range("A8").value = 'STT'
    for index, line in enumerate(list_stt, start=9):
        sht[f"A{index}"].number_format = "@"  
        sht[f"A{index}"].value = line   
    #print(list_stt) #A8

    list_hh = []
    img2 = img[720:1654,273:1035]
    text2 = pytesseract.image_to_string(img2, lang = 'eng',config= '--oem 3 --psm 6')
    lines2 = text2.strip().splitlines()
    filtered_lines = [line for line in lines2 if line.strip() != "" and line.strip() != " "]
    for index, line in enumerate(filtered_lines):
        if line == '(Product Name)':
            i = index
    while i >= 0:
        del filtered_lines[0]
        i-=1
    list_hh = (filtered_lines[:count])
    drink_elements = list_hh
    drink_elements = [element.replace("Sprite Plastic 300ml - Sprite chai nhua 300ml", "Đồ uống SPRITE 300ML 4X6 PET CARTON 2.0")
                    .replace("Dasani Water","Đồ uống DASANI 510ML 24 PET CARTON")
                    .replace("Fanta Plastic 300ml - Fanta chai nhua 300ml","Đồ uống FANTA ORANGE 300ML 4X6 PET CARTON")
                    .replace("Coke Plastic 300ml - Coca chai nhua 300ml","Đồ uống COKE 300ML 4X6 PET CARTON") for element in drink_elements] 
    sht.range("B8").value = 'Tên hàng hóa, dịch vụ'
    for index, line in enumerate(drink_elements, start=9):  
        sht[f"B{index}"].value = line 
    
    list_dvt = []
    img3 = img[720:1654, 1037:1187]
    text3 = pytesseract.image_to_string(img3, lang = 'eng',config= '--oem 3 --psm 6')
    lines3 = text3.strip().splitlines()
    filtered_lines = [line for line in lines3 if line.strip() != "" and line.strip() != " "]

    for index, line in enumerate(filtered_lines):
        if line == '(UoM)':
            i = index
    while i >= 0:
        del filtered_lines[0]
        i-=1
    list_dvt = [element.replace("Bottle","Két") for element in filtered_lines[:count]]
    sht.range("C8").value = 'Đơn vị tính'
    for index, line in enumerate(list_dvt, start=9):  
        sht[f"C{index}"].value = line 

    list_sl = []
    img4 = img[720:1654, 1190:1300]
    text4 = pytesseract.image_to_string(img4, lang = 'eng',config= '--oem 3 --psm 6')
    lines4 = text4.strip().splitlines()
    filtered_lines = [line for line in lines4 if line.strip() != "" and line.strip() != " "]

    for index, line in enumerate(filtered_lines):
        if line == '(Qty)':
            i = index
    while i >= 0:
        del filtered_lines[0]
        i-=1
    list_sl = [str(int(item) // 24) for item in filtered_lines[:count]]
    sht.range("D8").value = 'Số lượng'
    for index, line in enumerate(list_sl, start=9):  
        sht[f"D{index}"].number_format = "@"
        sht[f"D{index}"].value = line 

    list_dg = []
    img5 = img[720:1654, 1313:1462]
    text5 = pytesseract.image_to_string(img5, lang = 'eng',config= '--oem 3 --psm 6')
    lines5 = text5.strip().splitlines()
    filtered_lines = [item.replace(',', '') for item in lines5]
    for index, line in enumerate(filtered_lines):
        if line == '(Price Unit)':
            i = index
    while i >= 0:
        del filtered_lines[0]
        i-=1
    dg = [str(float(item)* 24) for item in filtered_lines[:count]]
    list_dg = [item.split('.')[0] for item in dg]
    #Đưa đơn giá về dạng number
    sht.range("E8").value = 'Đơn giá' 
    for index, line in enumerate(list_dg, start=9):  
        sht[f"E{index}"].number_format = "#,##0"
        sht[f"E{index}"].value = line

    list_tt = []
    img6 = img[720:1654, 1500:1664]
    text6 = pytesseract.image_to_string(img6, lang = 'eng',config= '--oem 3 --psm 6')
    lines6 = text6.strip().splitlines()
    filtered_lines = [line for line in lines6 if line.strip() != "" and line.strip() != " "]
    for index, line in enumerate(filtered_lines):
        if line == '(Amount)':
            i = index
    while i >= 0:
        del filtered_lines[0]
        i-=1
    list_tt = filtered_lines[:count]
    sht.range("F8").value = 'Thành tiền'
    for index, line in enumerate(list_tt, start=9):  
        sht[f"F{index}"].number_format = "#,##0"
        text_value = line
        numeric_value = float(text_value.replace('.', '').replace(',', ''))
        sht[f"F{index}"].value = numeric_value

    sum  = (filtered_lines[count:count+3])
    for i, item in enumerate(sum):
        if i == 0:
            sht.range("E4").value = 'Cộng tiền hàng'
            sht.range("F4").number_format = "#,##0"
            value = sum[i].replace(',', '.')
            numeric_value = float(value.replace('.', '').replace(',', ''))
            sht.range("F4").value = numeric_value
        if i == 1:
            sht.range("E5").value = 'Tiền thuế GTGT'
            sht.range("F5").number_format = "#,##0"
            value = sum[i].replace(',', '.')
            numeric_value = float(value.replace('.', '').replace(',', ''))
            sht.range("F5").value = numeric_value
        if i == 2:
            sht.range("E6").value = 'Tổng tiền thanh toán'
            sht.range("F6").number_format = "#,##0"
            value = sum[i].replace(',', '.')
            numeric_value = float(value.replace('.', '').replace(',', ''))
            sht.range("F6").value = numeric_value
    
    file_name = f"{no}_Coca_PO.xlsx"
    wb.save(fr'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Data_PO\{file_name}')
    excel_file_path = fr'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Data_PO\{file_name}'  
    make_columns_bold(excel_file_path, 1,6,1)
    make_columns_bold(excel_file_path, 1,6,5)
    make_rows_bold(excel_file_path, 1, 6, 8)
    wb = xw.Book(excel_file_path)
    change_font_size_and_save(wb, 12)
    process_all_worksheets(excel_file_path)
    
def main():
    word_po_path = "H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Customize PDF\word_po"
    pdf_folder_path = "H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\File_PDF_PO"
    for filename in os.listdir(word_po_path):
        if filename.endswith(".docx") or filename.endswith(".doc"):
            # Tạo đường dẫn đầy đủ cho tệp Word và tệp PDF
            word_path = os.path.join(word_po_path, filename)
            pdf_path = os.path.join(pdf_folder_path, os.path.splitext(filename)[0] + ".pdf")
            # Chuyển đổi từ Word sang PDF
            convert_word_to_pdf_single(word_path, pdf_path)
    pdf_to_png_PO()
    path_PO = 'H:\APP UNIVERSITY\CODE PYTHON\OpenCV-Master-Computer-Vision-in-Python\SourcecodeOCR\Source_Compare_PO'
    myList_PO = os.listdir(path_PO)
    for index, i in enumerate(myList_PO):
        img_path = os.path.join(path_PO, myList_PO[index])
        img = cv2.imread(img_path)
        img = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        img1 = img[15:180,430:1615]
        string1 =  pytesseract.image_to_string(img1, lang = 'vie',config= '--oem 3 --psm 6 preserve_interword_spaces=1')
        lines1 = string1.strip().splitlines()
        for element in lines1:
            if "Công ty" in element:
                words = element.split()
                if 'ty:' in words:
                    index_of_colon = words.index('ty:')
                    consumer = ' '.join(words[index_of_colon + 1:])   
        img2 = img[305:740,50:1800]
        string2 =  pytesseract.image_to_string(img2, lang = 'vie',config= '--oem 3 --psm 6 preserve_interword_spaces=1')
        lines2 = string2.strip().splitlines()
        data2 = []
        for index, line in enumerate(lines2):
            parts = line.split(':')
            if len(parts) > 1:
                content_after_colon = parts[1].strip()
                data2.append(content_after_colon)
        # #Số phiếu      
        no = data2[0]
        # #Vendor
        vendor = data2[1]
        # #Địa chỉ1
        address1 = data2[2]
        # #Tên cửa hàng
        ten_ch = data2[3]
        # #Địa chỉ 2
        address2 = data2[5]
        # #Note
        if len(data2) == 6:
            notes = ""
        else:
            notes = data2[6]
        img3 = img[1000:1654,50:430]
        string3 =  pytesseract.image_to_string(img3, lang = 'vie',config= '--oem 3 --psm 6 preserve_interword_spaces=1')
        lines3 = string3.strip().splitlines()
        data3 = []
        for index, line in enumerate(lines3):
            parts = line.split(':')
            if len(parts) > 1:
                content_after_colon = parts[1].strip()
                data3.append(content_after_colon)
        ngay_giaohang = data3[1]
        PO_cocacola(img,vendor,address1,consumer,ten_ch,address2,ngay_giaohang,no,notes)
        os.remove(img_path)
    Excel_Jusified_PO()
    print("Quá trình hoàn tất")

if __name__ == "__main__":
    main() 


